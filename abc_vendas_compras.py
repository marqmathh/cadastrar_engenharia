import tkinter as tk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side
import os
import pyautogui

def criar_janela_vendas():

    url = 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - 2021 - 2024.xlsx'
    df = pd.read_excel(url)

    dados_2021 = pd.read_excel(url, sheet_name='2021')
    dados_2022 = pd.read_excel(url, sheet_name='2022')
    dados_2023 = pd.read_excel(url, sheet_name='2023')
    dados_2024 = pd.read_excel(url, sheet_name='2024')
    dados_geral = pd.read_excel(url, sheet_name='Geral')

    def criar_2021():
        total_custo = sum(dados_2021['Valor total'])
        dados_2021['80 - 20 (%)'] = (dados_2021['Valor total'] / total_custo)
        df_final = dados_2021.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Vendas_2021.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_2022():
        total_custo = sum(dados_2022['Valor total'])
        dados_2022['80 - 20 (%)'] = (dados_2022['Valor total'] / total_custo)
        df_final = dados_2022.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Vendas_2022.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_2023():
        total_custo = sum(dados_2023['Valor total'])
        dados_2023['80 - 20 (%)'] = (dados_2023['Valor total'] / total_custo)
        df_final = dados_2023.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Vendas_2023.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_2024():
        total_custo = sum(dados_2024['Valor total'])
        dados_2024['80 - 20 (%)'] = (dados_2024['Valor total'] / total_custo)
        df_final = dados_2024.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Vendas_2024.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_geral():
        df_final = dados_geral.sort_values(by='Grupo', ascending=True)
        output_file = 'ABC_Vendas_geral.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        thick_border = Border(right=Side(style="thick"))
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor 2021" or col[0].value == "Valor 2022" or col[0].value == "Valor 2023" or col[0].value == "Valor 2024":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024" or col[0].value == "Grupo":
                for cell in col:
                    cell.border = thick_border
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def instrucao():
        pyautogui.alert("planilhas são salvas em Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas")

    def criar_interface():
        janela = tk.Tk()
        janela.title("Vendas - ABC")
        janela.geometry("400x400")
        
        tk.Label(janela, text="Selecione o periodo desejado:", font=("Arial", 14)).pack(pady=10)

        tk.Button(janela, text="Instrução", command=instrucao, width=30, bg="purple", fg="white").pack(pady=5)
        tk.Button(janela, text="2021", command=criar_2021, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="2022", command=criar_2022, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="2023", command=criar_2023, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="2024", command=criar_2024, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="Resumo - Geral", command=criar_geral, width=30, bg="yellow", fg="black").pack(pady=5)

        janela.mainloop()

    criar_interface()

def criar_janela_fornecedores():
    url_01 = 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - fornecedores.xlsx'
    df_forn = pd.read_excel(url_01)

    fornecedores_2021 = pd.read_excel(url_01, sheet_name='Fornecedores_2021')
    fornecedores_2022 = pd.read_excel(url_01, sheet_name='Fornecedores_2022')
    fornecedores_2023 = pd.read_excel(url_01, sheet_name='Fornecedores_2023')
    fornecedores_2024 = pd.read_excel(url_01, sheet_name='Fornecedores_2024')
    dados_geral_fornecedores = pd.read_excel(url_01, sheet_name='Geral_fornecedores')

    def criar_fornecedores_2021():
        total_custo = sum(fornecedores_2021['Valor total'])
        fornecedores_2021['80 - 20 (%)'] = (fornecedores_2021['Valor total'] / total_custo)
        df_final = fornecedores_2021.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Fornecedores_2021.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Nome do fornecedor do pedido de compra":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_fornecedores_2022():
        total_custo = sum(fornecedores_2022['Valor total'])
        fornecedores_2022['80 - 20 (%)'] = (fornecedores_2022['Valor total'] / total_custo)
        df_final = fornecedores_2022.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Fornecedores_2022.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Nome do fornecedor do pedido de compra":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_fornecedores_2023():
        total_custo = sum(fornecedores_2023['Valor total'])
        fornecedores_2023['80 - 20 (%)'] = (fornecedores_2023['Valor total'] / total_custo)
        df_final = fornecedores_2023.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Fornecedores_2023.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Nome do fornecedor do pedido de compra":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_fornecedores_2024():
        total_custo = sum(fornecedores_2024['Valor total'])
        fornecedores_2024['80 - 20 (%)'] = (fornecedores_2024['Valor total'] / total_custo)
        df_final = fornecedores_2024.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Fornecedores_2024.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Nome do fornecedor do pedido de compra":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_geral_fornecedores():
        df_final_01 = dados_geral_fornecedores.sort_values(by='Fornecedor', ascending=True)
        output_file = 'ABC_Fonecedores_geral.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final_01.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        thick_border = Border(right=Side(style="thick"))
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor 2021" or col[0].value == "Valor 2022" or col[0].value == "Valor 2023" or col[0].value == "Valor 2024":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024" or col[0].value == "Fornecedor":
                for cell in col:
                    cell.border = thick_border
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def instrucao_01():
        pyautogui.alert("planilhas são salvas em Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores")

    def criar_interface_01():
        janela = tk.Tk()
        janela.title("Fornecedores - ABC")
        janela.geometry("400x400")
        
        tk.Label(janela, text="Selecione o periodo desejado:", font=("Arial", 14)).pack(pady=10)

        tk.Button(janela, text="Instrução", command=instrucao_01, width=30, bg="purple", fg="white").pack(pady=5)
        tk.Button(janela, text="2021", command=criar_fornecedores_2021, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="2022", command=criar_fornecedores_2022, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="2023", command=criar_fornecedores_2023, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="2024", command=criar_fornecedores_2024, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="Resumo - Geral", command=criar_geral_fornecedores, width=30, bg="yellow", fg="black").pack(pady=5)

        janela.mainloop()
    
    criar_interface_01()

def criar_janela_compras():
    url_02 = 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - fornecedores.xlsx'
    df_com = pd.read_excel(url_02)

    compras_2021 = pd.read_excel(url_02, sheet_name='Compras_2021')
    compras_2022 = pd.read_excel(url_02, sheet_name='Compras_2022')
    compras_2023 = pd.read_excel(url_02, sheet_name='Compras_2023')
    compras_2024 = pd.read_excel(url_02, sheet_name='Compras_2024')
    dados_geral_compras = pd.read_excel(url_02, sheet_name='Compras_geral')

    def criar_compras_2021():
        total_custo = sum(compras_2021['Valor total'])
        compras_2021['80 - 20 (%)'] = (compras_2021['Valor total'] / total_custo)
        df_final =compras_2021.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Compras_2021.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto do item de pedido de compra":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_compras_2022():
        total_custo = sum(compras_2022['Valor total'])
        compras_2022['80 - 20 (%)'] = (compras_2022['Valor total'] / total_custo)
        df_final =compras_2022.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Compras_2022.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto do item de pedido de compra":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_compras_2023():
        total_custo = sum(compras_2023['Valor total'])
        compras_2023['80 - 20 (%)'] = (compras_2023['Valor total'] / total_custo)
        df_final =compras_2023.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Compras_2023.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto do item de pedido de compra":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_compras_2024():
        total_custo = sum(compras_2024['Valor total'])
        compras_2024['80 - 20 (%)'] = (compras_2024['Valor total'] / total_custo)
        df_final =compras_2024.sort_values(by='Valor total', ascending=False)
        output_file = 'ABC_Compras_2024.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto do item de pedido de compra":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_geral_compras():
        df_final_01 = dados_geral_compras.sort_values(by='Grupo', ascending=True)
        output_file = 'ABC_Compras_geral.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final_01.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        thick_border = Border(right=Side(style="thick"))
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Valor 2021" or col[0].value == "Valor 2022" or col[0].value == "Valor 2023" or col[0].value == "Valor 2024":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024" or col[0].value == "Grupo":
                for cell in col:
                    cell.border = thick_border
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def instrucao_01():
        pyautogui.alert("planilhas são salvas em Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras")

    def criar_interface_01():
        janela = tk.Tk()
        janela.title("Compras - ABC")
        janela.geometry("400x400")
        
        tk.Label(janela, text="Selecione o periodo desejado:", font=("Arial", 14)).pack(pady=10)

        tk.Button(janela, text="Instrução", command=instrucao_01, width=30, bg="purple", fg="white").pack(pady=5)
        tk.Button(janela, text="2021", command=criar_compras_2021, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="2022", command=criar_compras_2022, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="2023", command=criar_compras_2023, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="2024", command=criar_compras_2024, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="Resumo - Geral", command=criar_geral_compras, width=30, bg="yellow", fg="black").pack(pady=5)

        janela.mainloop()
    
    criar_interface_01()

def criar_janela_ordens():
    url_03 = 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - Ordens.xlsx'
    df_op = pd.read_excel(url_03)

    ordens_2023 = pd.read_excel(url_03, sheet_name='2023')
    ordens_2024 = pd.read_excel(url_03, sheet_name='2024')
    dados_geral_ordens = pd.read_excel(url_03, sheet_name='Geral')

    def criar_ordens_2023():
        df_final =ordens_2023.sort_values(by='Qtde.Total', ascending=False)
        output_file = 'ABC_Ordens_2023.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\ordens'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "DESCRIÇÃO DO PRODUTO":
                ws.column_dimensions[col_letter].width = 50
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_ordens_2024():
        df_final =ordens_2024.sort_values(by='Qtde.Total', ascending=False)
        output_file = 'ABC_Ordens_2024.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\ordens'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "DESCRIÇÃO DO PRODUTO":
                ws.column_dimensions[col_letter].width = 50
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def criar_ordens_geral():
        df_final_01 = dados_geral_ordens.sort_values(by='CÓDIGO DO PRODUTO', ascending=True)
        output_file = 'ABC_Ordens_geral.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\ordens'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_final_01.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        thick_border = Border(right=Side(style="thick"))
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Ordens 2023" or col[0].value == "CÓDIGO DO PRODUTO":
                for cell in col:
                    cell.border = thick_border
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def instrucao():
        pyautogui.alert("planilhas são salvas em Z:\ISO 9000 - SGQ\9 - PPCP\abc\ordens")

    def criar_interface_02():
        janela = tk.Tk()
        janela.title("Ordens - ABC")
        janela.geometry("400x400")
        
        tk.Label(janela, text="Selecione o periodo desejado:", font=("Arial", 14)).pack(pady=10)

        tk.Button(janela, text="Instrução", command=instrucao, width=30, bg="purple", fg="white").pack(pady=5)
        tk.Button(janela, text="2023", command=criar_ordens_2023, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="2024", command=criar_ordens_2024, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="Resumo - Geral", command=criar_ordens_geral, width=30, bg="yellow", fg="black").pack(pady=5)

        janela.mainloop()

    criar_interface_02()


app = tk.Tk()
app.title("ABC")
app.geometry("400x400")

# Botões da tela inicial
tk.Label(app, text="Selecione o calculo desejado:", font=("Arial", 14)).pack(pady=10)

tk.Button(app, text="ABC - Vendas", command=criar_janela_vendas, width=20, bg="green", fg="white").pack(pady=10)
tk.Button(app, text="ABC - Compras", command=criar_janela_compras, width=20, bg="purple", fg="white").pack(pady=10)
tk.Button(app, text="ABC - Fornecedores", command=criar_janela_fornecedores, width=20, bg="blue", fg="white").pack(pady=10)
tk.Button(app, text="ABC - OP's", command=criar_janela_ordens, width=20, bg="yellow", fg="black").pack(pady=10)

tk.Button(app, text="Sair", command=app.quit, bg="red", fg="white", width=20).pack(pady=10)

app.mainloop()