import tkinter as tk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import pyautogui

url = 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Setores de estoque.xlsx'
df = pd.read_excel(url)

def estoque_uso_e_consumo():
    filtro = df['Nome do setor de estoque'] == 'ESTOQUE USO E CONSUMO'
    df_filtrado = df[filtro]
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    output_file = 'Estoque uso e consumo.xlsx'
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def estoque_processo():
    filtro = df['Nome do setor de estoque'] == 'ESTOQUE PROCESSO'
    df_filtrado = df[filtro]
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    output_file = 'Estoque processo.xlsx'
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def estoque_em_terceiros():
    filtro = df['Nome do setor de estoque'] == 'ESTOQUE EM TERCEIROS'
    df_filtrado = df[filtro]
    output_file = 'Estoque em terceiros.xlsx'
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def estoque_retono_terceiros():
    filtro = df['Nome do setor de estoque'] == 'ESTOQUE RETORNO TERCEIROS'
    df_filtrado = df[filtro]
    output_file = 'Estoque retorno terceiros.xlsx'
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def estoque_de_clientes():
    filtro = df['Nome do setor de estoque'] == 'ESTOQUE DE CLIENTES'
    df_filtrado = df[filtro]
    output_file = 'Estoque de clientes.xlsx'
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def estoque_almoxarifado_central():
    filtro = df['Nome do setor de estoque'] == 'ESTOQUE ALMOXARIFADO CENTRAL'
    df_filtrado = df[filtro]
    output_file = 'Estoque almoxarifado central.xlsx'
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def estoque_expedicao_acabados():
    filtro = df['Nome do setor de estoque'] == 'ESTOQUE EXPEDIÇÃO ACABADOS'
    df_filtrado = df[filtro]
    output_file = 'Estoque expedição acabados.xlsx'
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def ativos_imobilizados():
    filtro = df['Nome do setor de estoque'] == 'ATIVOS IMOBILIZADOS'
    df_filtrado = df[filtro]
    output_file = 'Ativos imobilizados.xlsx'
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def almoxarifado_PHD():
    filtro = df['Nome do setor de estoque'] == 'ALMOXARIFADO PHD'
    df_filtrado = df[filtro]
    output_file = 'Almoxarifado PHD.xlsx'
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def estoque_processo_PHD():
    filtro = df['Nome do setor de estoque'] == 'EST. PROCESSO PHD'
    df_filtrado = df[filtro]
    output_file = 'Estoque processo PHD.xlsx'
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def acabados_PHD():
    filtro = df['Nome do setor de estoque'] == 'ACABADOS PHD'
    df_filtrado = df[filtro]
    output_file = 'Acabados PHD.xlsx'
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def estoque_almoxarifado_pre_producao():
    filtro = df['Nome do setor de estoque'] == 'ESTOQUE ALMOXARIFADO PRÉ-PRODUÇÃO'
    df_filtrado = df[filtro]
    output_file = 'Estoque almoxarifado pré-produção.xlsx'
    total_custo = sum(df_filtrado['Custo médio total'])
    df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
    df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
    pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
    caminho_completo = os.path.join(pasta_destino, output_file)
    df_filtrado.to_excel(caminho_completo, index=False)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    tabela = Table(displayName="Tabela", ref=ws.dimensions)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '#,##0.00' 
        if col[0].value == "80 - 20 (%)":
            for cell in col:
                if cell.row != 1: 
                    cell.number_format = '0.00%' 
        if col[0].value == "Descrição do produto":
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 
    wb.save(caminho_completo)

def instrucao():
    pyautogui.alert("planilhas são salvas em Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques")

def criar_interface():
    janela = tk.Tk()
    janela.title("Estoques")
    janela.geometry("400x550")
      
    tk.Label(janela, text="Selecione o estoque desejado:", font=("Arial", 14)).pack(pady=10)

    tk.Button(janela, text="Instrução", command=instrucao, width=30, bg="purple", fg="white").pack(pady=5)
    tk.Button(janela, text="Estoque almoxarifado central", command=estoque_almoxarifado_central, width=30, bg="green", fg="white").pack(pady=5)
    tk.Button(janela, text="Estoque almoxarifado pré-produção", command=estoque_almoxarifado_pre_producao, width=30, bg="green", fg="white").pack(pady=5)
    tk.Button(janela, text="Estoque expedição acabados", command=estoque_expedicao_acabados, width=30, bg="green", fg="white").pack(pady=5)
    tk.Button(janela, text="Estoque Processo", command=estoque_processo, width=30, bg="yellow", fg="black").pack(pady=5)
    tk.Button(janela, text="Ativo imobilizado", command=ativos_imobilizados, width=30, bg="yellow", fg="black").pack(pady=5)
    tk.Button(janela, text="Estoque uso e consumo", command=estoque_uso_e_consumo, width=30, bg="yellow", fg="black").pack(pady=5)
    tk.Button(janela, text="Estoque em terceiros", command=estoque_em_terceiros, width=30, bg="blue", fg="white").pack(pady=5)
    tk.Button(janela, text="Estoque retorno terceiros", command=estoque_retono_terceiros, width=30, bg="blue", fg="white").pack(pady=5)
    tk.Button(janela, text="Estoque de clientes", command=estoque_de_clientes, width=30, bg="blue", fg="white").pack(pady=5)
    tk.Button(janela, text="Almoxarifado PHD", command=almoxarifado_PHD, width=30, bg="red", fg="white").pack(pady=5)
    tk.Button(janela, text="Estoque processo PHD", command=estoque_processo_PHD, width=30, bg="red", fg="white").pack(pady=5)
    tk.Button(janela, text="Acabados_PHD", command=acabados_PHD, width=30, bg="red", fg="white").pack(pady=5)

    janela.mainloop()

criar_interface()
