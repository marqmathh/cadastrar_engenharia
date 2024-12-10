from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import tkinter as tk
from tkinter import messagebox, ttk
import os, shutil, win32com.client, time, ctypes, pyautogui, math, requests, patoolib
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side
from sqlalchemy import create_engine, Column, Integer, String, Float
from sqlalchemy.orm import sessionmaker, declarative_base

def atualizar_planilhas_versao_2():
    # Função para mover os arquivos para uma pasta específica
    def mover_arquivos(pasta_origem, pasta_destino):
        for item in os.listdir(pasta_origem):
            item_path = os.path.join(pasta_origem, item)
            try:
                if os.path.isfile(item_path):
                    shutil.move(item_path, pasta_destino)
            except Exception as e:
                print(f'Erro ao mover {item_path} para {pasta_destino}: {e}')
    # Função para excluir os arquivos em uma pasta destino
    def excluir_itens(pasta):
        for item in os.listdir(pasta):
            item_path = os.path.join(pasta, item)
            try:
                if os.path.isfile(item_path):
                    os.remove(item_path) 
                elif os.path.isdir(item_path):
                    shutil.rmtree(item_path) 
            except Exception as e:
                print(f'Erro ao excluir {item_path}: {e}')
    # Função para pressionar os botões de exportar
    def exportar():
        time.sleep(30)
        export_button = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ZDBExportButton"]')))
        export_button.click()
        try: 
            embed_export_button = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="EmbedExportXLSMenuItem"]')))
            time.sleep(1)
            embed_export_button.click()
            
            final_button = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[32]/div/div[2]/div/footer/div/button[1]')))
            time.sleep(1)
            final_button.click()
            print('Planilha exportada com sucesso (modelo padrão)')
            time.sleep(1)
        except Exception:
            embed_export_button = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SumAndPvtEmbedExportXLSMenuItem"]')))
            time.sleep(1)
            embed_export_button.click()
            
            final_button = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[34]/div/div[2]/div/footer/div/button[1]')))
            time.sleep(1)
            final_button.click()
            time.sleep(1)
            print('Planilha exportada com sucesso (modelo secundário)')
    # Função para abrir uma nova guia no nomus
    def mudar_Link(texto,i):
        navegador.execute_script("window.open('');")
        navegador.switch_to.window(navegador.window_handles[i])
        navegador.get(texto)
    # Função para exportar as planilhas do nomus
    def exporta_Nomus():
        urls = [
            ("https://reports.nomus.com.br/open-view/751489003498986462", 0),  # PC - Follow up
            ("https://reports.nomus.com.br/open-view/751489003495455916", 1),  # SC COMPLETA
            ("https://reports.nomus.com.br/open-view/751489003499869977", 2),  # Cotação
            ("https://reports.nomus.com.br/open-view/751489003499311067", 3),  # NF Tipo de movimentação
            ("https://reports.nomus.com.br/open-view/751489003497642590", 4),  # Fornecedores - TB-06
            ("https://reports.nomus.com.br/open-view/751489003385051213", 5),  # Ordens Lead
            ("https://reports.nomus.com.br/open-view/751489003418257074", 6),  # Necessidade
            ("https://reports.nomus.com.br/open-view/751489003348744953", 7),  # ID-04e
            ("https://reports.nomus.com.br/open-view/751489003473916593", 8),  # Tempo de produção
            ("https://reports.nomus.com.br/open-view/751489003396077915", 9),  # Entradas
            ("https://reports.nomus.com.br/open-view/751489003343390887", 10), # Pedido de compra
            ("https://reports.nomus.com.br/open-view/751489003375648746", 11), # Pv's
            ("https://reports.nomus.com.br/open-view/751489003343401232", 12), # SC's
            ("https://reports.nomus.com.br/open-view/751489003343402198", 13), # Entradas compras
            ("https://reports.nomus.com.br/open-view/751489003383877015", 14), # Faturamento periodo
            ("https://reports.nomus.com.br/open-view/751489003384381506", 15), # CRM
            ("https://reports.nomus.com.br/open-view/751489003381705320", 16), # Vendas
            ("https://reports.nomus.com.br/open-view/751489003396836362", 17), # Emissão de PV
            ("https://reports.nomus.com.br/open-view/751489003384833393", 18), # Status do PV
            ("https://reports.nomus.com.br/open-view/751489003501209517", 19), # Familia
            ("https://reports.nomus.com.br/open-view/751489003299922759", 20), # CONTROLE DA PRODUÇÃO
            ("https://reports.nomus.com.br/open-view/751489003399788047", 21), # ENTRADAS (LIVRO)
            ("https://reports.nomus.com.br/open-view/751489003415814176", 22), # Nfs emitidas
            ("https://reports.nomus.com.br/open-view/751489003399436517", 23), # Saída
            ("https://reports.nomus.com.br/open-view/751489003406446494", 24), # Dw Compras
            ("https://reports.nomus.com.br/open-view/751489003342815143", 25), # Modelo NF's
            ("https://reports.nomus.com.br/open-view/751489003415906765", 26), # Grupos
            ("https://reports.nomus.com.br/open-view/751489003432511966", 27), # movimentação - pvs
            ("https://reports.nomus.com.br/open-view/751489003502954549", 28), # Descrições
            ("https://reports.nomus.com.br/open-view/751489003502944448", 29), # Estoque - tempo real
            ("https://reports.nomus.com.br/open-view/751489003521923921", 30), # qtde_estoque
            ("https://reports.nomus.com.br/open-view/751489003423136168", 31), # ABC Vendas
            ("https://reports.nomus.com.br/open-view/751489003526023430", 32), # Inventários
        ]
        navegador.get(urls[0][0])
        time.sleep(10)
        exportar() 
        for url, index in urls[1:]:
            mudar_Link(url, index)
            exportar() 
        time.sleep(60)
        navegador.quit()
    # Função para salvar os arquivos do download em outro lugar antes de começar a atualização
    def salva_Arquivos():
        pasta_destino = r'Z:\PUBLICO\## Salva arquivos'
        pasta_downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
        mover_arquivos(pasta_downloads, pasta_destino)
    # Função para mover os arquivos baixados para a pasta do servidor
    def move_Arquivos():
        pasta_destino = r'Z:\PUBLICO\### Banco de dados'
        pasta_downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
        excluir_itens(pasta_destino)
        mover_arquivos(pasta_downloads, pasta_destino)
    # Função para voltar os arquivos para a pasta Download
    def voltar_Arquivos():
        pasta_destino = r'Z:\PUBLICO\## Salva arquivos'
        pasta_downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
        mover_arquivos(pasta_destino, pasta_downloads)
    # Função para atulizar as planilhas padrões
    def atualizar_Planilhas(local_Planilha):
        excel = win32com.client.DispatchEx('Excel.Application')
        excel.Visible = True
        wb = excel.Workbooks.Open(local_Planilha)
        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone() 
        wb.Close(SaveChanges=True)
        print(f"Planilha atualizada com sucesso: {local_Planilha}")
        excel.Quit()
    # Função de atualizar as planilhas que não possuimos acesso
    def atualizar_Planilhas_especiais(local_Planilha):
        excel = win32com.client.DispatchEx('Excel.Application')
        excel.Visible = True
        wb = excel.Workbooks.Open(local_Planilha)
        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone() 
        data_atual = datetime.now().strftime("%Y-%m-%d")
        nome_arquivo, extensao = os.path.splitext(local_Planilha)
        novo_nome = f"{nome_arquivo} - {data_atual}{extensao}"
        wb.SaveAs(novo_nome)
        wb.Close(SaveChanges=True)
        excel.Quit()
        
        print(f"Planilha atualizada e salva como: {novo_nome}")
    # Atualiza as planilhas
    def atualizar_Excel():
        planilhas = [
            # 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Lead time - Familia.xlsx',
            'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Monitoramento de estoque.xlsx',
            'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\OP_porcentagem.xlsx',
            'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Vendas.xlsx',
            'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Indicadores.xlsx',
            'Z:\\ISO 9000 - SGQ\\9 - PPCP\\.PCP\\Controle\\Controle.xlsm',
            'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Setores de estoque.xlsx'
        ]
        planilhas_especiais = [
            # 'Z:\\ISO 9000 - SGQ\\6-PROCESSO SUPRIMENTOS\\REGISTROS\\TB-06_AvalProvedoresExternos-Rev05.xlsx',
            'Z:\\PUBLICO\\Araujo\\Planilhas-Indicadores\\Gráficos-IDs-PCP-2024 - IDs 04a 04b 13b.xlsx',
            'Z:\\PUBLICO\\Araujo\\Planilhas-Indicadores\\Gráfico-ID-Produção-2024 - ID-02-01-10-2024.xlsx',
            'Z:\\PUBLICO\\Araujo\\Planilhas-Indicadores\\Graficos-IDs-Compras-2024 - IDs 09 13a 13c.xlsx',
            'Z:\\PUBLICO\\Araujo\\Planilhas-Indicadores\\Graficos-IDs-Engenharia-2024 - IDs10a 10b 11 12 e 14.xlsx',
            'Z:\\ISO 9000 - SGQ\\5-PROCESSO PROJETOS\\REGISTROS\\Controle - AT.xlsm'
        ]
        for planilha in planilhas:
            atualizar_Planilhas(planilha)
        for planilha in planilhas_especiais:
            atualizar_Planilhas_especiais(planilha)
    # Exibe as intruções
    def instrucoes():
        messagebox.showinfo("Instrução de Atualização de Indicadores e Planilhas", "Olá\n\nEste programa é responsável pela atualização dos indicadores e das planilhas padrão.\n\nPlanilhas que são atualizadas automaticamente:\nLead time\nMonitoramento de estoque\nSetores de estoque\nTB-06\nControle\nControle AT\nIndicadores salvos no Publico/Araujo\n\nNo entanto, existem três planilhas que, devido à interdependência de suas atualizações, não podem ser automatizadas. Caso seja necessário utilizar a Curva ABC, será preciso atualizar manualmente as seguintes planilhas:\n\n'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - 2021 - 2024.xlsx'\n\n'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - fornecedores.xlsx'\n\n'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - Ordens.xlsx'\n\nPara atualizar, basta entrar nela e pressionar em 'Atualizar Tudo'")
    # Exibe a mensagem de inicialização do programa
    def inicializacao():
        messagebox.showinfo("Inicialização", "A atividade terá início em breve. Por favor, não utilize o computador até que a atualização seja concluída.")
    # Exibe a mensagem de finalização do programa
    def finalizacao():
        messagebox.showinfo("Finalização", "Atividade concluída com sucesso!")
    # Atualiza tudo
    def atualizacao_geral():
        inicializacao()
        time.sleep(0.5)
        salva_Arquivos()
        time.sleep(0.5)
        exporta_Nomus()
        time.sleep(5)
        move_Arquivos()
        time.sleep(0.5)
        voltar_Arquivos()
        time.sleep(0.5)
        atualizar_Excel()
        time.sleep(0.5)
    # Atualiza tudo e avisa
    def atualizacao_nivel_1():
        atualizacao_geral()
        finalizacao()
    # Atualiza tudo e bloqueia
    def atualizacao_nivel_2():
        atualizacao_geral()
        ctypes.windll.user32.LockWorkStation()
    # Atualiza tudo e suspende
    def atualizacao_nivel_3():
        atualizacao_geral()
        os.system("rundll32.exe powrprof.dll,SetSuspendState 0,1,0")
    # Atualiza tudo e desliga
    def atualizacao_nivel_4():
        atualizacao_geral()
        os.system("shutdown /s /t 10")
    # Cria a interface da aplicação
    def criar_interface():
        janela = tk.Tk()
        janela.title("Automação de Processos")
        janela.geometry("320x450")
        janela.config(bg="lightblue")  
        tk.Button(janela, text="Instruções", command=instrucoes, width=30, bg="blue", fg="white").pack(pady=(20,5))
        tk.Button(janela, text="Atualizar tudo", command=atualizacao_nivel_1, width=30, bg="purple1", fg="white").pack(pady=5)
        tk.Button(janela, text="Atualizar tudo e bloquear", command=atualizacao_nivel_2, width=30, bg="purple2", fg="white").pack(pady=5)
        tk.Button(janela, text="Atualizar tudo e suspender", command=atualizacao_nivel_3, width=30, bg="purple3", fg="white").pack(pady=5)
        tk.Button(janela, text="Atualizar tudo e desligar", command=atualizacao_nivel_4, width=30, bg="purple4", fg="white").pack(pady=5)
        tk.Button(janela, text="1° - Salvar arquivos", command=salva_Arquivos, width=30, bg="purple4", fg="white").pack(pady=5)
        tk.Button(janela, text="2° - Baixar planilhas", command=exporta_Nomus, width=30, bg="purple4", fg="white").pack(pady=5)
        tk.Button(janela, text="3° - Mover arquivos", command=move_Arquivos, width=30, bg="purple4", fg="white").pack(pady=5)
        tk.Button(janela, text="4° - Voltar Arquivos", command=voltar_Arquivos, width=30, bg="purple4", fg="white").pack(pady=5)
        tk.Button(janela, text="5° - Atualizar planilhas", command=atualizar_Excel, width=30, bg="purple4", fg="white").pack(pady=5)
        janela.mainloop()
    # Rodar a aplicação
    servico = Service(ChromeDriverManager().install())
    navegador = webdriver.Chrome(service=servico)
    criar_interface()

def atualizar_nomus_versao_1():
    arquivo_excel = r'Z:\PUBLICO\Araujo\Cadastros\Cadastros.xlsx'
    arquivo_excel_01 = r'Z:\PUBLICO\Araujo\Cadastros\edicao.xlsx'

    # Cadastrar o item
    def criar_janela_cadastro_produtos():
        df = pd.read_excel(arquivo_excel, sheet_name='Cadastro')

        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico)

        usuario = entry_usuario.get() 
        senha = entry_senha.get()
        origem = '0 - Nacional (exceto as indicadas nos códigos de 3 a 5)'
        hoje = datetime.today()
        dia = hoje.strftime("%d")
        mes = hoje.strftime("%m")
        ano = hoje.year
        dia_atual = f'{dia}{mes}{ano}'
        empresa_tspro = '//*[@id="empresa_1"]'
        empresa_phd = '//*[@id="empresa_3"]'
        estoque_processo = '//*[@id="setor_2"]'
        estoque_em_terceiros = '//*[@id="setor_5"]'
        esoque_retorno_terceiros = '//*[@id="setor_7"]'
        estoque_de_clientes = '//*[@id="setor_8"]'
        estoque_almoxarifado_central = '//*[@id="setor_9"]'
        estoque_expedicao_acabados = '//*[@id="setor_10"]'
        ativo_imobilizados = '//*[@id="setor_11"]'
        estoque_almoxarifado_pre_producao = '//*[@id="setor_16"]'
        estoque_uso_e_consumo = '//*[@id="setor_17"]'
        almoxarifado_phd = '//*[@id="setor_13"]'
        est_processo_phd = '//*[@id="setor_14"]'
        acabados_phd = '//*[@id="setor_15"]'

        index = 0

        def login(usuario, senha):
                tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
                tela_usuario.send_keys(usuario)
                tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
                tela_senha.send_keys(senha)    
                tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_produtos():
            navegador.get('https://tspro.nomus.com.br/tspro/Produto.do?metodo=Pesquisar') # Produtos

        def criar_produtos():
            criar_produtos_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_criar_produto"]'))) # Criar produtos
            criar_produtos_01.click()

        def cadastrar_aba_inicial(cod, descricao, und, tipo, grupo, metodo):
            criar_produtos_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="nome"]'))) # Código do produto
            criar_produtos_02.send_keys(cod)

            criar_produtos_03 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//table/tbody/tr[7]/td[2]/textarea'))) # Descrição do produto
            criar_produtos_03.send_keys(descricao)


            unidade_medida_select = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="idUnidadeMedida"]'))) # Unidade de medida
            select_unidade_medida = Select(unidade_medida_select)
            select_unidade_medida.select_by_visible_text(und)

            tipo_produto = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tipoProduto_id"]'))) # Tipo de produto
            select_tipo_produto = Select(tipo_produto)
            select_tipo_produto.select_by_visible_text(tipo)

            grupo_produto = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="nomeGrupoProduto_id_select"]'))) # Grupo de produto
            select_grupo_produto = Select(grupo_produto)
            select_grupo_produto.select_by_visible_text(grupo)

            metodo_produto = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="idPadraoSuprimento"]'))) # Método de suprimento
            select_metodo_produto = Select(metodo_produto)
            select_metodo_produto.select_by_visible_text(metodo)

        def cadastrar_empresas():
            empresa = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-28"]'))) # Empresas
            empresa.click()

            if tipo == 'EMBALAGENS':
                empresa_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_uso_e_consumo))) # Seleciona o estoque
                empresa_02.click()

            if tipo == 'INSUMOS E MATERIAS PRIMAS':
                empresa_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_processo))) # Seleciona o estoque
                empresa_02.click()

                empresa_03 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_em_terceiros))) # Seleciona o estoque
                empresa_03.click()

                empresa_04 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, esoque_retorno_terceiros))) # Seleciona o estoque
                empresa_04.click()

                empresa_05 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_almoxarifado_central))) # Seleciona o estoque
                empresa_05.click()

                empresa_06 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_almoxarifado_pre_producao))) # Seleciona o estoque
                empresa_06.click()

            if tipo == 'MATERIAL DE TERCEIROS':
                empresa_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_almoxarifado_pre_producao))) # Seleciona o estoque
                empresa_02.click()

                empresa_03 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_de_clientes))) # Seleciona o estoque
                empresa_03.click()

            if tipo == 'MATERIAL DE USO E CONSUMO' or tipo == 'USO E CONSUMO COM ESTOQUE':
                empresa_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_uso_e_consumo))) # Seleciona o estoque
                empresa_02.click()

            if tipo == 'PRODUTO ACABADO':
                empresa_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_expedicao_acabados))) # Seleciona o estoque
                empresa_02.click()

            if tipo == 'PRODUTO SEMI ACABADO':
                empresa_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_processo))) # Seleciona o estoque
                empresa_02.click()

                empresa_03 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_almoxarifado_central))) # Seleciona o estoque
                empresa_03.click()

                empresa_04 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_almoxarifado_pre_producao))) # Seleciona o estoque
                empresa_04.click()
            
            if tipo == 'SERVICOS':
                empresa_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, estoque_processo))) # Seleciona o estoque
                empresa_02.click()

                empresa_03 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, esoque_retorno_terceiros))) # Seleciona o estoque
                empresa_03.click()

            if tipo == 'ATIVO IMOBILIZADO':
                empresa_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, ativo_imobilizados))) # Seleciona o estoque
                empresa_02.click()

        def cadastrar_fiscal(ncm_excel, descricao_fiscal):
            fiscal = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-29"]'))) # Fiscal
            fiscal.click()

            fiscal_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="modificarDescricaoNFe"]'))) # Fiscal
            fiscal_01.click()

            fiscal_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="descricaoNFe"]'))) # Descrição fiscal
            fiscal_02.send_keys(descricao_fiscal)
            
            fiscal_04_select = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//table/tbody/tr[9]/td[2]/select'))) # origem do produto
            select_fiscal_04 = Select(fiscal_04_select)
            select_fiscal_04.select_by_visible_text(origem)

            ncm_escolher = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="id_nomeNcm"]'))) # Descrição fiscal
            ncm_escolher.send_keys(ncm_excel)
            time.sleep(1)
            ncm_escolher.send_keys(Keys.RETURN) 
            time.sleep(0.5)       

        def cadastrar_pcp(pcp_ressuprimento):
            pcp_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-34"]'))) # PCP
            pcp_01.click()

            pcp_01_select = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_politica_ressuprimento"]/td[2]/select'))) # PCP ressuprimento
            select_pcp_01 = Select(pcp_01_select)
            select_pcp_01.select_by_visible_text(pcp_ressuprimento)

            fabrica = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-36"]'))) # FABRICA
            fabrica.click()

            fabrica_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="idSugereQtdeProduzidaApontamento"]'))) # FABRICA
            fabrica_01.click()

            fabrica_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="idGeraReporteProducaoAutomatico"]'))) # FABRICA
            fabrica_02.click()

        def cadastrar_mrp():
            mrp = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-35"]'))) # MRP
            mrp.click()

            mrp_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_lote_multiplo"]/td[2]/input'))) # lote multiplo
            mrp_01.send_keys(0)

            mrp_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_lote_minimo"]/td[2]/input'))) # lote minimo
            mrp_02.send_keys(0)

            mrp_03 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_lote_maximo"]/td[2]/input'))) # lote maximo
            mrp_03.send_keys(0)

            mrp_04 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_estoque_de_seguranca"]/td[2]/input'))) # estoque de seguraça
            mrp_04.send_keys(0)

            mrp_05 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_estoque_maximo"]/td[2]/input'))) # estoque maximo
            mrp_05.send_keys(0)

        def salvar_produto():
            salvar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) # Salvar cadastro
            salvar.click()

        def cadastrar_custo(custo_padrao, dia_atual):
            if metodo == 'Comprado':
                custo = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-37"]'))) # custo
                custo.click()

                custo_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="custoPadraoCompra_id"]'))) # custo padrao
                custo_01.send_keys(custo_padrao)

                custo_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="dataReferenciaCusto_id"]'))) # data referencia
                custo_02.send_keys(dia_atual)

                ressup_prod = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="atualizaCustoPadraoCompra_id"]'))) # Método de suprimento
                select_ressup_prod = Select(ressup_prod)
                select_ressup_prod.select_by_visible_text('Atualiza custo padrão de compra com base no custo de reposição')

                salvar_produto()
            else:
                salvar_produto()

        def voltar_no_item(cod, financeiro_excel):
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input"))) # Código do produto
            produto_01.send_keys(cod)

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{cod}']"))) #Seleciona o elemento
            botao_buscar.click()

            xpaths = ['//*[@id="produtoAtivoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoAtivoLiberado_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoLiberado_itemSubMenu_editarProduto"]']
            for xpath in xpaths:
                try:
                    botao_entrar_editar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    botao_entrar_editar.click()
                    break
                except Exception:
                    pass 

            fiscal = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-29"]'))) # Fiscal
            fiscal.click()

            classe_financeira_selecionar_select = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="id_idClassificacaoFinanceiraPadrao"]'))) # Unidade de medida
            select_classe_financeira_selecionar = Select(classe_financeira_selecionar_select)
            select_classe_financeira_selecionar.select_by_visible_text(financeiro_excel)

            salvar_produto()

        def cadastrar_produto(ncm_excel, cod, descricao, und, tipo, grupo, metodo, descricao_fiscal, pcp_ressuprimento, custo_padrao, dia_atual, financeiro_excel):
            criar_produtos()
            cadastrar_aba_inicial(cod, descricao, und, tipo, grupo, metodo)
            cadastrar_empresas()
            cadastrar_pcp(pcp_ressuprimento)
            cadastrar_mrp()
            cadastrar_fiscal(ncm_excel, descricao_fiscal)
            cadastrar_custo(custo_padrao, dia_atual)
            voltar_no_item(cod, financeiro_excel)
                
        entrar_nomus()
        tela_produtos()
        
        for index, row in df.iterrows():
            cod = row['Código']
            descricao = row['Descrição']
            und = row['Unidade']
            tipo = row['Tipo do produto']
            grupo = row['Grupo do produto']
            metodo = row['Metodo']
            descricao_fiscal = row['Descrição Fiscal']
            ncm_excel = row['NCM']
            financeiro_excel = row['Classificacao']
            pcp_ressuprimento = row['PCP Ressuprimento']
            custo_padrao = row['Custo padrao']
                        
            cadastrar_produto(ncm_excel, cod, descricao, und, tipo, grupo, metodo, descricao_fiscal, pcp_ressuprimento, custo_padrao, dia_atual, financeiro_excel)

    # Adicionar LM ao produto
    def criar_janela_lista_de_materiais():
        df = pd.read_excel(arquivo_excel, sheet_name='LM')
        
        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico) 

        usuario = entry_usuario.get() 
        senha = entry_senha.get()
        cod = df.iloc[1, 0]
            
        index = 0

        def login(usuario, senha):
            tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
            tela_usuario.send_keys(usuario)
            tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
            tela_senha.send_keys(senha)    
            tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_produtos():
            navegador.get('https://tspro.nomus.com.br/tspro/Produto.do?metodo=Pesquisar') # Produtos

        def entra_produto(cod):
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input"))) # Código do produto
            produto_01.send_keys(cod)

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{cod}']"))) #Seleciona o elemento
            botao_buscar.click()

            xpaths = ['//*[@id="produtoAtivoAguardandoLiberacao_itemSubMenu_acessarListaMateriais"]', '//*[@id="produtoAtivoLiberado_itemSubMenu_acessarListaMateriais"]']
            for xpath in xpaths:
                try:
                    botao_entrar_lm = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    botao_entrar_lm.click()
                    break
                except Exception:
                    pass    

        def criar_lm(cod):
            campo = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="geral"]/table/tbody/tr[1]/td[2]/input')))

            valor = campo.get_attribute("value")

            if valor: 
                produto_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/table/tbody/tr[1]/td[2]/input'))) # Código do produto
                produto_02.send_keys(f' - lista - 1')

                botao_criar_nova_lista = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar_como_nova_lista"]'))) # Buscar produtos
                botao_criar_nova_lista.click()

                botao_criar_nova_lista_05 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="marcaredesmarcar"]'))) # Buscar produtos
                botao_criar_nova_lista_05.click()

                botao_criar_nova_lista_06 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_Acoes"]'))) # Buscar produtos
                botao_criar_nova_lista_06.click()

                botao_criar_nova_lista_07 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_botao.excluir.componentes"]'))) # Buscar produtos
                botao_criar_nova_lista_07.click()




            else: 
                produto_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/table/tbody/tr[1]/td[2]/input'))) # Código do produto
                produto_02.send_keys(cod)

                botao = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]')))
                botao.click()

        def cadastrar_lm(item_lista, qtde_item_lista, natureza, posicao):
            botao_adicionar_estrutura = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_acessaradicionaritemestrutura"]'))) # Buscar produtos
            botao_adicionar_estrutura.click()

            botao_adicionar_estrutura = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[1]/td[2]/a/i"))) # Buscar produtos
            botao_adicionar_estrutura.click()

            procurar_item = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="id_nomeProdutoPesquisa"]'))) # Código do produto
            procurar_item.send_keys(item_lista)
            
            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            time.sleep(2)

            botao_adicionar_estrutura_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '(//*[@id="idProdutoSelecionado"])[1]'))) # Buscar produtos
            botao_adicionar_estrutura_01.click()

            botao_finalizar_estrutura_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="selecionar"]'))) # Buscar produtos
            botao_finalizar_estrutura_01.click()

            qtde_necessaria_item = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="qtdeNecessaria_id"]'))) # Código do produto
            qtde_necessaria_item.send_keys(qtde_item_lista)

            natureza_produto = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="naturezaConsumo"]'))) # Grupo de produto
            select_natureza_produto = Select(natureza_produto)
            select_natureza_produto.select_by_visible_text(natureza)

            botao_posicao_do_item = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[14]/td[2]/input"))) # Buscar produtos
            botao_posicao_do_item.send_keys(posicao)
            
            botao_criar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) # Buscar produtos
            botao_criar.click()
        
        entrar_nomus()
        tela_produtos()
        entra_produto(cod)
        criar_lm(cod)
        
        time.sleep(5)

        for index, row in df.iterrows():
            item_lista = row['Itens LM']
            qtde_item_lista = row['QTDE']
            natureza = row['Natureza']
            posicao = row['Posicao']
                                
            cadastrar_lm(item_lista, qtde_item_lista, natureza, posicao)

    # Adicionar RO ao produto
    def criar_janela_rp():
        df = pd.read_excel(arquivo_excel, sheet_name='RP')
        
        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico) 

        usuario = entry_usuario.get() 
        senha = entry_senha.get()

        index = 0

        def login(usuario, senha):
            tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
            tela_usuario.send_keys(usuario)
            tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
            tela_senha.send_keys(senha)    
            tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_produtos():
            navegador.get('https://tspro.nomus.com.br/tspro/Produto.do?metodo=Pesquisar') # Produtos

        def acessar_produto(cod):
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input"))) # Código do produto
            produto_01.clear()
            produto_01.send_keys(cod)

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{cod}']"))) #Seleciona o elemento
            botao_buscar.click()

            xpaths = ['//*[@id="produtoAtivoAguardandoLiberacao_itemSubMenu_acessarRoteiroProduto"]', '//*[@id="produtoAtivoLiberado_itemSubMenu_acessarRoteiroProduto"]']
            for xpath in xpaths:
                try:
                    botao_entrar_lm = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    botao_entrar_lm.click()
                    break
                except Exception:
                    pass 

        def adicionar_rp_no_produto(rp_padrao):
            rp_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//table/tbody/tr[1]/td[2]/textarea')))
            rp_01.clear()
            rp_01.send_keys(rp_padrao)
            try:
                botao_rp_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar_como_roteiro_adicional"]'))) 
                botao_rp_02.click()
            except:
                botao_rp_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) 
                botao_rp_02.click()

        def exportar_rp_para_os_itens(rp_padrao, cod):
            navegador.get('https://tspro.nomus.com.br/tspro/GrupoRoteiroProducao.do?metodo=pesquisarPaginado')

            export_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input")))
            export_01.clear()
            export_01.send_keys(rp_padrao)

            botao_export_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisarpaginado"]'))) 
            botao_export_02.click()

            botao_exxxx = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{rp_padrao}']"))) #Seleciona o elemento
            botao_exxxx.click()

            botao_export_0222 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="grupoRoteiroProducaoSubmenu_itemSubMenu_erro.selecionar.GrupoRoteiroProducao"]'))) 
            botao_export_0222.click()

            botao_export_02222 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-11"]'))) 
            botao_export_02222.click()

            produto_AA = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input"))) # Código do produto
            produto_AA.clear()
            produto_AA.send_keys(cod)

            botao_send_export_02222 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) 
            botao_send_export_02222.click()

            time.sleep(1)

            botao_export_03 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="marcaredesmarcar"]'))) 
            botao_export_03.click()

            botao_export_99 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_vincularprodutos"]'))) 
            botao_export_99.click()

            time.sleep(1)

            salvar_rp_geral = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) 
            salvar_rp_geral.click()

            export_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input")))
            export_01.send_keys(rp_padrao)

            botao_export_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisarpaginado"]'))) 
            botao_export_02.click()

            botao_export_03 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="marcaredesmarcar"]'))) 
            botao_export_03.click()

            botao_export_04 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_Acoes"]'))) 
            botao_export_04.click()

            botao_export_05 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_exportar_roteiro_produto_template"]'))) 
            botao_export_05.click()

            export_06 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="idComportamentoExportacao"]'))) # Tipo de produto
            select_export_06 = Select(export_06)
            select_export_06.select_by_visible_text('Exportar roteiro excluindo todos os roteiros do produto destino')

            botao_export_07 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="salvar"]'))) 
            botao_export_07.click()

            time.sleep(10)

        entrar_nomus()   

        for index, row in df.iterrows():
            cod = row['Código']
            rp_padrao = row['rot padrao']
                        
            tela_produtos()
            acessar_produto(cod)
            adicionar_rp_no_produto(rp_padrao)
            exportar_rp_para_os_itens(rp_padrao,cod)

    # Abrir planilha de cadastro de itens
    def criar_janela_planilha_cadastro():
        os.startfile(arquivo_excel)

    # Abrir planilha de edição de itens
    def criar_janela_planilha_edicao():
        os.startfile(arquivo_excel_01)

    # Função editar aba inicial
    def editar_itens_geral():
        df = pd.read_excel(arquivo_excel_01, sheet_name='Principal')

        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico)

        usuario = entry_usuario.get() 
        senha = entry_senha.get()

        index = 0

        def login(usuario, senha):
                tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
                tela_usuario.send_keys(usuario)
                tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
                tela_senha.send_keys(senha)    
                tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_produtos():
            navegador.get('https://tspro.nomus.com.br/tspro/Produto.do?metodo=Pesquisar') # Produtos

        def acessar_produto(cod):
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input")))
            produto_01.clear()
            produto_01.send_keys(cod)

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{cod}']"))) #Seleciona o elemento
            botao_buscar.click()

            xpaths = ['//*[@id="produtoAtivoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoAtivoLiberado_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoLiberado_itemSubMenu_editarProduto"]']
            for xpath in xpaths:
                try:
                    botao_entrar_editar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    botao_entrar_editar.click()
                    break
                except Exception:
                    pass 
        
        def edicao_geral(cod_editado,descricao,und,tipo,grupo,metodo,atv_desat):
            criar_produtos_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="nome"]')))
            criar_produtos_02.clear()
            criar_produtos_02.send_keys(cod_editado)

            criar_produtos_03 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//table/tbody/tr[7]/td[2]/textarea'))) # Descrição do produto
            criar_produtos_03.clear()
            criar_produtos_03.send_keys(descricao)

            unidade_medida_select = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="idUnidadeMedida"]'))) # Unidade de medida
            select_unidade_medida = Select(unidade_medida_select)
            select_unidade_medida.select_by_visible_text(und)

            tipo_produto = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tipoProduto_id"]'))) # Tipo de produto
            select_tipo_produto = Select(tipo_produto)
            select_tipo_produto.select_by_visible_text(tipo)

            grupo_produto = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="nomeGrupoProduto_id_select"]'))) # Grupo de produto
            select_grupo_produto = Select(grupo_produto)
            select_grupo_produto.select_by_visible_text(grupo)

            metodo_produto = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="idPadraoSuprimento"]'))) # Método de suprimento
            select_metodo_produto = Select(metodo_produto)
            select_metodo_produto.select_by_visible_text(metodo)

            if atv_desat == 'Sim':
                botao_ativar_desativar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ativoId"]'))) #Seleciona o elemento
                botao_ativar_desativar.click()

                salvar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) # Salvar cadastro
                salvar.click()
            else:
                salvar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) # Salvar cadastro
                salvar.click()


        entrar_nomus()
        
        for index, row in df.iterrows():
            cod = row['Código']
            atv_desat = row['Ativar/Desativar']
            cod_editado = row['Novo Código']
            descricao = row['Descrição']
            und = row['Unidade']
            tipo = row['Tipo do produto']
            grupo = row['Grupo do produto']
            metodo = row['Metodo']
                        
            tela_produtos()
            acessar_produto(cod)
            edicao_geral(cod_editado,descricao,und,tipo,grupo,metodo,atv_desat)

    # Função editar aba empresas
    def editar_empresas():
        df = pd.read_excel(arquivo_excel_01, sheet_name='Empresas')

        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico)

        usuario = entry_usuario.get() 
        senha = entry_senha.get()

        index = 0

        def login(usuario, senha):
                tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
                tela_usuario.send_keys(usuario)
                tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
                tela_senha.send_keys(senha)    
                tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_produtos():
            navegador.get('https://tspro.nomus.com.br/tspro/Produto.do?metodo=Pesquisar') # Produtos

        def acessar_produto(cod, empresa_excel):
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input")))
            produto_01.clear()
            produto_01.send_keys(cod)

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{cod}']"))) #Seleciona o elemento
            botao_buscar.click()

            xpaths = ['//*[@id="produtoAtivoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoAtivoLiberado_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoLiberado_itemSubMenu_editarProduto"]']
            for xpath in xpaths:
                try:
                    botao_entrar_editar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    botao_entrar_editar.click()
                    break
                except Exception:
                    pass 
            
            empresa = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-28"]'))) # Empresas
            empresa.click()

            empresa_selecao = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, empresa_excel)))
            empresa_selecao.click()
        
            salvar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) # Salvar cadastro
            salvar.click()

        entrar_nomus()
        
        for index, row in df.iterrows():
            cod = row['Código']
            empresa_excel = row['Coluna1']
                        
            tela_produtos()
            acessar_produto(cod, empresa_excel)

    # Função editar aba fiscal
    def editar_fiscal():
        df = pd.read_excel(arquivo_excel_01, sheet_name='Fiscal')

        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico)

        usuario = entry_usuario.get() 
        senha = entry_senha.get()

        index = 0

        def login(usuario, senha):
                tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
                tela_usuario.send_keys(usuario)
                tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
                tela_senha.send_keys(senha)    
                tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_produtos():
            navegador.get('https://tspro.nomus.com.br/tspro/Produto.do?metodo=Pesquisar') # Produtos

        def acessar_produto(cod):
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input")))
            produto_01.clear()
            produto_01.send_keys(cod)

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{cod}']"))) #Seleciona o elemento
            botao_buscar.click()

            xpaths = ['//*[@id="produtoAtivoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoAtivoLiberado_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoLiberado_itemSubMenu_editarProduto"]']
            for xpath in xpaths:
                try:
                    botao_entrar_editar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    botao_entrar_editar.click()
                    break
                except Exception:
                    pass

        def edicao_fiscal_aa(descricao_fiscal, origem, ncm_excel, financeiro_excel):
            fiscal = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-29"]'))) # Fiscal
            fiscal.click()

            fiscal_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="descricaoNFe"]'))) # Descrição fiscal
            fiscal_02.clear()
            fiscal_02.send_keys(descricao_fiscal)
            
            fiscal_04_select = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//table/tbody/tr[9]/td[2]/select'))) # origem do produto
            select_fiscal_04 = Select(fiscal_04_select)
            select_fiscal_04.select_by_visible_text(origem)

            ncm_escolher = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="id_nomeNcm"]'))) # Descrição fiscal
            ncm_escolher.clear()
            ncm_escolher.send_keys(ncm_excel)
            time.sleep(1)
            ncm_escolher.send_keys(Keys.RETURN) 
            time.sleep(0.5)   

            classe_financeira_selecionar_select = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="id_idClassificacaoFinanceiraPadrao"]'))) # Unidade de medida
            select_classe_financeira_selecionar = Select(classe_financeira_selecionar_select)
            select_classe_financeira_selecionar.select_by_visible_text(financeiro_excel)

            salvar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) # Salvar cadastro
            salvar.click()

        entrar_nomus()
        
        for index, row in df.iterrows():
            cod = row['Código']
            origem = row['Origem']
            descricao_fiscal = row['Descrição Fiscal']
            ncm_excel = row['NCM']
            financeiro_excel = row['Classificacao']
                        
            tela_produtos()
            acessar_produto(cod)
            edicao_fiscal_aa(descricao_fiscal, origem, ncm_excel, financeiro_excel)

    # Função editar aba PCP
    def editar_pcp():
        df = pd.read_excel(arquivo_excel_01, sheet_name='PCP')

        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico)

        usuario = entry_usuario.get() 
        senha = entry_senha.get()

        index = 0

        def login(usuario, senha):
                tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
                tela_usuario.send_keys(usuario)
                tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
                tela_senha.send_keys(senha)    
                tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_produtos():
            navegador.get('https://tspro.nomus.com.br/tspro/Produto.do?metodo=Pesquisar') # Produtos

        def acessar_produto(cod):
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input")))
            produto_01.clear()
            produto_01.send_keys(cod)

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{cod}']"))) #Seleciona o elemento
            botao_buscar.click()

            xpaths = ['//*[@id="produtoAtivoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoAtivoLiberado_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoLiberado_itemSubMenu_editarProduto"]']
            for xpath in xpaths:
                try:
                    botao_entrar_editar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    botao_entrar_editar.click()
                    break
                except Exception:
                    pass

        def editar_pcp_aa(pcp_ressuprimento):
            pcp_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-34"]'))) # PCP
            pcp_01.click()

            pcp_01_select = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_politica_ressuprimento"]/td[2]/select'))) # PCP ressuprimento
            select_pcp_01 = Select(pcp_01_select)
            select_pcp_01.select_by_visible_text(pcp_ressuprimento)

            salvar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) # Salvar cadastro
            salvar.click()
        
        entrar_nomus()

        for index, row in df.iterrows():
            cod = row['Código']
            pcp_ressuprimento = row['PCP Ressuprimento']
            
            tela_produtos()
            acessar_produto(cod)
            editar_pcp_aa(pcp_ressuprimento)

    # Função editar aba MRP
    def editar_mrp():
        df = pd.read_excel(arquivo_excel_01, sheet_name='MRP')

        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico)

        usuario = entry_usuario.get() 
        senha = entry_senha.get()
        pcp_ressuprimento = 'Ressuprimento para estoque através do plano de produção e MRP'

        index = 0
        
        def login(usuario, senha):
                tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
                tela_usuario.send_keys(usuario)
                tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
                tela_senha.send_keys(senha)    
                tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_produtos():
            navegador.get('https://tspro.nomus.com.br/tspro/Produto.do?metodo=Pesquisar') # Produtos

        def acessar_produto(cod):
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input")))
            produto_01.clear()
            produto_01.send_keys(cod)

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{cod}']"))) #Seleciona o elemento
            botao_buscar.click()

            xpaths = ['//*[@id="produtoAtivoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoAtivoLiberado_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoLiberado_itemSubMenu_editarProduto"]']
            for xpath in xpaths:
                try:
                    botao_entrar_editar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    botao_entrar_editar.click()
                    break
                except Exception:
                    pass

        def editar_mrp_aa(pcp_ressuprimento, multiplo, minimo, maximo, seguranca, est_maximo):
            pcp_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-34"]'))) # PCP
            pcp_01.click()

            pcp_01_select = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_politica_ressuprimento"]/td[2]/select'))) # PCP ressuprimento
            select_pcp_01 = Select(pcp_01_select)
            select_pcp_01.select_by_visible_text(pcp_ressuprimento)

            mrp = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-35"]'))) # MRP
            mrp.click()

            mrp_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_lote_multiplo"]/td[2]/input'))) # lote multiplo
            mrp_01.send_keys(multiplo)

            mrp_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_lote_minimo"]/td[2]/input'))) # lote minimo
            mrp_02.send_keys(minimo)

            mrp_03 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_lote_maximo"]/td[2]/input'))) # lote maximo
            mrp_03.send_keys(maximo)

            mrp_04 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_estoque_de_seguranca"]/td[2]/input'))) # estoque de seguraça
            mrp_04.send_keys(seguranca)

            mrp_05 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tr_estoque_maximo"]/td[2]/input'))) # estoque maximo
            mrp_05.send_keys(est_maximo)

            salvar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) # Salvar cadastro
            salvar.click()

        entrar_nomus()

        for index, row in df.iterrows():
            cod = row['Código']
            multiplo = row['multiplo']
            minimo = row['minimo']
            maximo = row['maximo']
            seguranca = row['est. Seguranca']
            est_maximo = row['est. Maximo']
            
            tela_produtos()
            acessar_produto(cod)
            editar_mrp_aa(pcp_ressuprimento, multiplo, minimo, maximo, seguranca, est_maximo)

    # Função editar aba custos
    def editar_custos():
        df = pd.read_excel(arquivo_excel_01, sheet_name='Custo')

        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico)

        usuario = entry_usuario.get() 
        senha = entry_senha.get()
        hoje = datetime.today()
        dia = hoje.strftime("%d")
        mes = hoje.strftime("%m")
        ano = hoje.year
        dia_atual = f'{dia}{mes}{ano}'

        index = 0

        def login(usuario, senha):
                tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
                tela_usuario.send_keys(usuario)
                tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
                tela_senha.send_keys(senha)    
                tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_produtos():
            navegador.get('https://tspro.nomus.com.br/tspro/Produto.do?metodo=Pesquisar') # Produtos

        def acessar_produto(cod):
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input")))
            produto_01.clear()
            produto_01.send_keys(cod)

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{cod}']"))) #Seleciona o elemento
            botao_buscar.click()

            xpaths = ['//*[@id="produtoAtivoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoAtivoLiberado_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoAguardandoLiberacao_itemSubMenu_editarProduto"]', '//*[@id="produtoInativoLiberado_itemSubMenu_editarProduto"]']
            for xpath in xpaths:
                try:
                    botao_entrar_editar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    botao_entrar_editar.click()
                    break
                except Exception:
                    pass

        def editar_custo(custo_padrao, dia_atual):
            metodo_produto = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="idPadraoSuprimento"]'))) # Método de suprimento
            select_metodo_produto = Select(metodo_produto)
            select_metodo_produto.select_by_visible_text('Comprado')

            custo = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ui-id-37"]'))) # custo
            custo.click()

            custo_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="custoPadraoCompra_id"]'))) # custo padrao
            custo_01.send_keys(custo_padrao)

            custo_02 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="dataReferenciaCusto_id"]'))) # data referencia
            custo_02.send_keys(dia_atual)

            ressup_prod = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="atualizaCustoPadraoCompra_id"]'))) # Método de suprimento
            select_ressup_prod = Select(ressup_prod)
            select_ressup_prod.select_by_visible_text('Atualiza custo padrão de compra com base no custo de reposição')

            salvar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_salvar"]'))) # Salvar cadastro
            salvar.click()

        entrar_nomus()

        for index, row in df.iterrows():
            cod = row['Código']
            custo_padrao = row['Custo Padrao']
            
            tela_produtos()
            acessar_produto(cod)
            editar_custo(custo_padrao, dia_atual)

    # Tela de edição de itens
    def criar_janela_menu_edicao():
        janela = tk.Tk()
        janela.title("Edição de itens")
        janela.geometry("400x450")
        janela.config(bg="lightblue")  
        tk.Label(janela, text="Qual aba será editada ?", font=("Arial", 14), bg="lightblue").pack(pady=10)

        tk.Button(janela, text="Abrir planilha", command=criar_janela_planilha_edicao, width=20, bg="RoyalBlue1", fg="black").pack(pady=10)
        tk.Button(janela, text="Geral", command=editar_itens_geral, width=20, bg="turquoise1", fg="black").pack(pady=10)
        tk.Button(janela, text="Empresas", command=editar_empresas, width=20, bg="turquoise2", fg="black").pack(pady=10)
        tk.Button(janela, text="Fiscal", command=editar_fiscal, width=20, bg="turquoise3", fg="black").pack(pady=10)
        tk.Button(janela, text="PCP", command=editar_pcp, width=20, bg="PaleTurquoise2", fg="black").pack(pady=10)
        tk.Button(janela, text="MRP", command=editar_mrp, width=20, bg="SkyBlue1", fg="black").pack(pady=10)
        tk.Button(janela, text="Custos", command=editar_custos, width=20, bg="SteelBlue1", fg="black").pack(pady=10)

        janela.mainloop()

    # Tela de cadastro de produtos
    def criar_janela_cadastro():
        janela_01 = tk.Tk()
        janela_01.title("Edição de itens")
        janela_01.geometry("400x400")
        janela_01.config(bg="lightblue")  
        tk.Label(janela_01, text="O que será cadastrado ?", font=("Arial", 14), bg="lightblue").pack(pady=10)

        tk.Button(janela_01, text="Abrir planilha", command=criar_janela_planilha_cadastro, width=20, bg="DarkOrange2", fg="black").pack(pady=10)
        tk.Button(janela_01, text="Cadastrar Produtos", command=criar_janela_cadastro_produtos, width=20, bg="green", fg="white").pack(pady=10)
        tk.Button(janela_01, text="Cadastrar LM", command=criar_janela_lista_de_materiais, width=20, bg="blue", fg="white").pack(pady=10)
        tk.Button(janela_01, text="Cadastrar RP", command=criar_janela_rp, width=20, bg="NavajoWhite4", fg="white").pack(pady=10)

        janela_01.mainloop()

    # Tela inicial do app
    app = tk.Tk()
    app.title("Engenharia")
    app.geometry("400x400")
    app.config(bg="lightblue")  

    tk.Label(app, text="Selecione a opção desejada: ", font=("Arial", 14), bg="lightblue").pack(pady=5)
    label_usuario = tk.Label(app, text="Usuário:", bg="lightblue")
    label_usuario.pack(padx=5, pady=5)

    entry_usuario = tk.Entry(app, bg="white", bd=2, relief="solid", justify="center")
    entry_usuario.pack(padx=5, pady=5)

    label_senha = tk.Label(app, text="Senha:",  bg="lightblue")
    label_senha.pack(padx=5, pady=5)

    entry_senha = tk.Entry(app, show="*", bg="white", bd=2, relief="solid", justify="center")
    entry_senha.pack(padx=5, pady=5)

    tk.Button(app, text="Cadastrar itens", command=criar_janela_cadastro, width=20, bg="gold", fg="black").pack(pady=10)
    tk.Button(app, text="Editar cadastros", command=criar_janela_menu_edicao, width=20, bg="goldenrod", fg="black").pack(pady=10)

    app.mainloop()

def abc_geral():

    def criar_janela_vendas():
        url = 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - 2021 - 2024.xlsx'
        df = pd.read_excel(url)

        dados_2021 = pd.read_excel(url, sheet_name='2021')
        dados_2022 = pd.read_excel(url, sheet_name='2022')
        dados_2023 = pd.read_excel(url, sheet_name='2023')
        dados_2024 = pd.read_excel(url, sheet_name='2024')
        dados_geral = pd.read_excel(url, sheet_name='Geral')

        def criar_2021():
            total_custo = sum(dados_2021['Valor total'])
            dados_2021['80 - 20 (%)'] = (dados_2021['Valor total'] / total_custo)
            df_final = dados_2021.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Vendas_2021.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Descrição":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_2022():
            total_custo = sum(dados_2022['Valor total'])
            dados_2022['80 - 20 (%)'] = (dados_2022['Valor total'] / total_custo)
            df_final = dados_2022.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Vendas_2022.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Descrição":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_2023():
            total_custo = sum(dados_2023['Valor total'])
            dados_2023['80 - 20 (%)'] = (dados_2023['Valor total'] / total_custo)
            df_final = dados_2023.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Vendas_2023.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Descrição":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_2024():
            total_custo = sum(dados_2024['Valor total'])
            dados_2024['80 - 20 (%)'] = (dados_2024['Valor total'] / total_custo)
            df_final = dados_2024.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Vendas_2024.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Descrição":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_geral():
            df_final = dados_geral.sort_values(by='Grupo', ascending=True)
            output_file = 'ABC_Vendas_geral.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            thick_border = Border(right=Side(style="thick"))
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor 2021" or col[0].value == "Valor 2022" or col[0].value == "Valor 2023" or col[0].value == "Valor 2024":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024" or col[0].value == "Grupo":
                    for cell in col:
                        cell.border = thick_border
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def instrucao():
            pyautogui.alert("planilhas são salvas em Z:\ISO 9000 - SGQ\9 - PPCP\abc\vendas")

        def criar_interface():
            janela = tk.Tk()
            janela.title("Vendas - ABC")
            janela.geometry("400x400")
            
            tk.Label(janela, text="Selecione o periodo desejado:", font=("Arial", 14)).pack(pady=10)

            tk.Button(janela, text="Instrução", command=instrucao, width=30, bg="purple", fg="white").pack(pady=5)
            tk.Button(janela, text="2021", command=criar_2021, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="2022", command=criar_2022, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="2023", command=criar_2023, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="2024", command=criar_2024, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="Resumo - Geral", command=criar_geral, width=30, bg="yellow", fg="black").pack(pady=5)

            janela.mainloop()

        criar_interface()

    def criar_janela_fornecedores():
        url_01 = 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - fornecedores.xlsx'
        df_forn = pd.read_excel(url_01)

        fornecedores_2021 = pd.read_excel(url_01, sheet_name='Fornecedores_2021')
        fornecedores_2022 = pd.read_excel(url_01, sheet_name='Fornecedores_2022')
        fornecedores_2023 = pd.read_excel(url_01, sheet_name='Fornecedores_2023')
        fornecedores_2024 = pd.read_excel(url_01, sheet_name='Fornecedores_2024')
        dados_geral_fornecedores = pd.read_excel(url_01, sheet_name='Geral_fornecedores')

        def criar_fornecedores_2021():
            total_custo = sum(fornecedores_2021['Valor total'])
            fornecedores_2021['80 - 20 (%)'] = (fornecedores_2021['Valor total'] / total_custo)
            df_final = fornecedores_2021.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Fornecedores_2021.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Nome do fornecedor do pedido de compra":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_fornecedores_2022():
            total_custo = sum(fornecedores_2022['Valor total'])
            fornecedores_2022['80 - 20 (%)'] = (fornecedores_2022['Valor total'] / total_custo)
            df_final = fornecedores_2022.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Fornecedores_2022.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Nome do fornecedor do pedido de compra":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_fornecedores_2023():
            total_custo = sum(fornecedores_2023['Valor total'])
            fornecedores_2023['80 - 20 (%)'] = (fornecedores_2023['Valor total'] / total_custo)
            df_final = fornecedores_2023.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Fornecedores_2023.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Nome do fornecedor do pedido de compra":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_fornecedores_2024():
            total_custo = sum(fornecedores_2024['Valor total'])
            fornecedores_2024['80 - 20 (%)'] = (fornecedores_2024['Valor total'] / total_custo)
            df_final = fornecedores_2024.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Fornecedores_2024.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Nome do fornecedor do pedido de compra":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_geral_fornecedores():
            df_final_01 = dados_geral_fornecedores.sort_values(by='Fornecedor', ascending=True)
            output_file = 'ABC_Fonecedores_geral.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final_01.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            thick_border = Border(right=Side(style="thick"))
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor 2021" or col[0].value == "Valor 2022" or col[0].value == "Valor 2023" or col[0].value == "Valor 2024":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024" or col[0].value == "Fornecedor":
                    for cell in col:
                        cell.border = thick_border
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def instrucao_01():
            pyautogui.alert("planilhas são salvas em Z:\ISO 9000 - SGQ\9 - PPCP\abc\fornecedores")

        def criar_interface_01():
            janela = tk.Tk()
            janela.title("Fornecedores - ABC")
            janela.geometry("400x400")
            
            tk.Label(janela, text="Selecione o periodo desejado:", font=("Arial", 14)).pack(pady=10)

            tk.Button(janela, text="Instrução", command=instrucao_01, width=30, bg="purple", fg="white").pack(pady=5)
            tk.Button(janela, text="2021", command=criar_fornecedores_2021, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="2022", command=criar_fornecedores_2022, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="2023", command=criar_fornecedores_2023, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="2024", command=criar_fornecedores_2024, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="Resumo - Geral", command=criar_geral_fornecedores, width=30, bg="yellow", fg="black").pack(pady=5)

            janela.mainloop()
        
        criar_interface_01()

    def criar_janela_compras():
        url_02 = 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - fornecedores.xlsx'
        df_com = pd.read_excel(url_02)

        compras_2021 = pd.read_excel(url_02, sheet_name='Compras_2021')
        compras_2022 = pd.read_excel(url_02, sheet_name='Compras_2022')
        compras_2023 = pd.read_excel(url_02, sheet_name='Compras_2023')
        compras_2024 = pd.read_excel(url_02, sheet_name='Compras_2024')
        dados_geral_compras = pd.read_excel(url_02, sheet_name='Compras_geral')

        def criar_compras_2021():
            total_custo = sum(compras_2021['Valor total'])
            compras_2021['80 - 20 (%)'] = (compras_2021['Valor total'] / total_custo)
            df_final =compras_2021.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Compras_2021.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Descrição do produto do item de pedido de compra":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_compras_2022():
            total_custo = sum(compras_2022['Valor total'])
            compras_2022['80 - 20 (%)'] = (compras_2022['Valor total'] / total_custo)
            df_final =compras_2022.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Compras_2022.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Descrição do produto do item de pedido de compra":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_compras_2023():
            total_custo = sum(compras_2023['Valor total'])
            compras_2023['80 - 20 (%)'] = (compras_2023['Valor total'] / total_custo)
            df_final =compras_2023.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Compras_2023.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Descrição do produto do item de pedido de compra":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_compras_2024():
            total_custo = sum(compras_2024['Valor total'])
            compras_2024['80 - 20 (%)'] = (compras_2024['Valor total'] / total_custo)
            df_final =compras_2024.sort_values(by='Valor total', ascending=False)
            output_file = 'ABC_Compras_2024.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor total":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "80 - 20 (%)":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "Descrição do produto do item de pedido de compra":
                    ws.column_dimensions[col_letter].width = 70
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_geral_compras():
            df_final_01 = dados_geral_compras.sort_values(by='Grupo', ascending=True)
            output_file = 'ABC_Compras_geral.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final_01.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            thick_border = Border(right=Side(style="thick"))
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Valor 2021" or col[0].value == "Valor 2022" or col[0].value == "Valor 2023" or col[0].value == "Valor 2024":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '#,##0.00' 
                if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024":
                    for cell in col:
                        if cell.row != 1: 
                            cell.number_format = '0.00%' 
                if col[0].value == "% Valor Total 2021" or col[0].value == "% Valor Total 2022" or col[0].value == "% Valor Total 2023" or col[0].value == "% Valor Total 2024" or col[0].value == "Grupo":
                    for cell in col:
                        cell.border = thick_border
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def instrucao_01():
            pyautogui.alert("planilhas são salvas em Z:\ISO 9000 - SGQ\9 - PPCP\abc\compras")

        def criar_interface_01():
            janela = tk.Tk()
            janela.title("Compras - ABC")
            janela.geometry("400x400")
            
            tk.Label(janela, text="Selecione o periodo desejado:", font=("Arial", 14)).pack(pady=10)

            tk.Button(janela, text="Instrução", command=instrucao_01, width=30, bg="purple", fg="white").pack(pady=5)
            tk.Button(janela, text="2021", command=criar_compras_2021, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="2022", command=criar_compras_2022, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="2023", command=criar_compras_2023, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="2024", command=criar_compras_2024, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="Resumo - Geral", command=criar_geral_compras, width=30, bg="yellow", fg="black").pack(pady=5)

            janela.mainloop()
        
        criar_interface_01()

    def criar_janela_ordens():
        url_03 = 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Curva ABC - Ordens.xlsx'
        df_op = pd.read_excel(url_03)

        ordens_2023 = pd.read_excel(url_03, sheet_name='2023')
        ordens_2024 = pd.read_excel(url_03, sheet_name='2024')
        dados_geral_ordens = pd.read_excel(url_03, sheet_name='Geral')

        def criar_ordens_2023():
            df_final =ordens_2023.sort_values(by='Qtde.Total', ascending=False)
            output_file = 'ABC_Ordens_2023.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\ordens'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "DESCRIÇÃO DO PRODUTO":
                    ws.column_dimensions[col_letter].width = 50
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_ordens_2024():
            df_final =ordens_2024.sort_values(by='Qtde.Total', ascending=False)
            output_file = 'ABC_Ordens_2024.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\ordens'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "DESCRIÇÃO DO PRODUTO":
                    ws.column_dimensions[col_letter].width = 50
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def criar_ordens_geral():
            df_final_01 = dados_geral_ordens.sort_values(by='CÓDIGO DO PRODUTO', ascending=True)
            output_file = 'ABC_Ordens_geral.xlsx'
            pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\abc\ordens'
            if not os.path.exists(pasta_destino):
                os.makedirs(pasta_destino)
            caminho_completo = os.path.join(pasta_destino, output_file)
            df_final_01.to_excel(caminho_completo, index=False)
            wb = load_workbook(caminho_completo)
            ws = wb.active
            tabela = Table(displayName="Tabela", ref=ws.dimensions)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
            thick_border = Border(right=Side(style="thick"))
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col[0].value == "Ordens 2023" or col[0].value == "CÓDIGO DO PRODUTO":
                    for cell in col:
                        cell.border = thick_border
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
            wb.save(caminho_completo)

        def instrucao():
            pyautogui.alert("planilhas são salvas em Z:\ISO 9000 - SGQ\9 - PPCP\abc\ordens")

        def criar_interface_02():
            janela = tk.Tk()
            janela.title("Ordens - ABC")
            janela.geometry("400x400")
            
            tk.Label(janela, text="Selecione o periodo desejado:", font=("Arial", 14)).pack(pady=10)

            tk.Button(janela, text="Instrução", command=instrucao, width=30, bg="purple", fg="white").pack(pady=5)
            tk.Button(janela, text="2023", command=criar_ordens_2023, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="2024", command=criar_ordens_2024, width=30, bg="green", fg="white").pack(pady=5)
            tk.Button(janela, text="Resumo - Geral", command=criar_ordens_geral, width=30, bg="yellow", fg="black").pack(pady=5)

            janela.mainloop()

        criar_interface_02()


    app = tk.Tk()
    app.title("ABC")
    app.geometry("400x400")

    # Botões da tela inicial
    tk.Label(app, text="Selecione o calculo desejado:", font=("Arial", 14)).pack(pady=10)

    tk.Button(app, text="ABC - Vendas", command=criar_janela_vendas, width=20, bg="green", fg="white").pack(pady=10)
    tk.Button(app, text="ABC - Compras", command=criar_janela_compras, width=20, bg="purple", fg="white").pack(pady=10)
    tk.Button(app, text="ABC - Fornecedores", command=criar_janela_fornecedores, width=20, bg="blue", fg="white").pack(pady=10)
    tk.Button(app, text="ABC - OP's", command=criar_janela_ordens, width=20, bg="yellow", fg="black").pack(pady=10)

    app.mainloop()

def calculos_geral():
    # Funções para abrir as janelas específicas
    def abrir_tabela_pesos():
        janela_tabela = tk.Toplevel(app)
        janela_tabela.title("Tabela de Pesos")
        janela_tabela.geometry("500x500")
        criar_janela_tabela_pesos(janela_tabela)

    def abrir_mangueiras():
        janela_mangueiras = tk.Toplevel(app)
        janela_mangueiras.title("Mangueiras")
        janela_mangueiras.geometry("500x650")
        criar_janela_mangueira(janela_mangueiras)

    # def abrir_bracos():
    #     janela_bracos = tk.Toplevel(app)
    #     janela_bracos.title("Braços")
    #     janela_bracos.geometry("700x900")
    #     criar_janela_bracos(janela_bracos)

    # def abrir_succao():
    #     janela_succao = tk.Toplevel(app)
    #     janela_succao.title("Sucção")
    #     janela_succao.geometry("400x300")
    #     tk.Label(janela_succao, text="Funcionalidade de Sucção em desenvolvimento.", pady=20).pack()

    # def abrir_perda_de_carga():
        # janela_succao = tk.Toplevel(app)
        # janela_succao.title("Perda de Carga")
        # janela_succao.geometry("400x300")
        # tk.Label(janela_succao, text="Funcionalidade de Perda de carga em desenvolvimento.", pady=20).pack()


    # Função para criar a janela de Tabela de Pesos
    def criar_janela_tabela_pesos(janela):
        # Dicionários para tipos de cálculo e materiais
        tipo_calculo = {
            "Tubo Redondo": "TR",
            "Tubo Quadrado/Retangular": "TQR",
            "Cantoneira": "CT",
            "Barra Redonda": "BR",
            "Barra Retangular/Quadrada": "BRQ",
            "Barra Sextavada": "BS"
        }

        material_aplicacao = {
            "Aço": 7.85,
            "Alumínio": 2.71,
            "Alumínio Fundido": 2.6,
            "Bronze": 8.8,
            "Latão": 8.55,
            "Chumbo": 11.31,
            "Chumbo fundido": 10.37,
            "Teflon": 2.2,
            "PVC": 1.45,
            "Policetal": 1.425,
            "Polietileno (PEAD)": 0.92,
            "Polipropileno": 0.91,
            "Nylon": 1.14,
            "HMW": 0.942,
        }

        def atualizar_campos(event):
            for widget in campos_frame.winfo_children():
                widget.destroy()
            
            calculo = tipo_calculo.get(tipo_combobox.get())

            if calculo == "TR":
                criar_campo("Diâmetro Externo (mm)", "di_ext")
                criar_campo("Espessura (mm)", "esp")
                criar_campo("Comprimento (mm)", "comp")
            elif calculo == "TQR":
                criar_campo("Lado 1 (mm)", "l1")
                criar_campo("Lado 2 (mm)", "l2")
                criar_campo("Espessura (mm)", "esp")
                criar_campo("Comprimento (mm)", "comp")
            elif calculo == "CT":
                criar_campo("Lado 1 (mm)", "l1")
                criar_campo("Lado 2 (mm)", "l2")
                criar_campo("Espessura (mm)", "esp")
                criar_campo("Comprimento (mm)", "comp")
            elif calculo == "BR":
                criar_campo("Diâmetro (mm)", "di")
                criar_campo("Comprimento (mm)", "comp")
            elif calculo == "BRQ":
                criar_campo("Lado 1 (mm)", "l1")
                criar_campo("Lado 2 (mm)", "l2")
                criar_campo("Comprimento (mm)", "comp")
            elif calculo == "BS":
                criar_campo("Fator A (mm)", "fa")
                criar_campo("Comprimento (mm)", "comp")

        def criar_campo(label_text, field_name):
            frame = tk.Frame(campos_frame)
            frame.pack(anchor="w", pady=2)
            tk.Label(frame, text=label_text, width=20, anchor="w").pack(side="left")
            entry = tk.Entry(frame, width=15)
            entry.pack(side="right")
            campos[field_name] = entry

        def converter_valor(valor):
            try:
                return float(valor.replace(",", "."))
            except ValueError:
                raise ValueError(f"Valor inválido: {valor}")

        def calcular():
            tipo_escolhido = tipo_combobox.get()
            material_escolhido = material_combobox.get()
            
            calculo = tipo_calculo.get(tipo_escolhido)
            material = material_aplicacao.get(material_escolhido)

            try:
                if calculo == "TR":
                    di_ext = converter_valor(campos["di_ext"].get())
                    esp = converter_valor(campos["esp"].get())
                    comp = converter_valor(campos["comp"].get())
                    kg_metro = ((di_ext**2) * (0.7854/1000) * material) - (((di_ext - (esp * 2))**2) * (0.7854/1000) * material)
                    kg_total = (kg_metro * comp) / 1000
                elif calculo == "TQR":
                    l1 = converter_valor(campos["l1"].get())
                    l2 = converter_valor(campos["l2"].get())
                    esp = converter_valor(campos["esp"].get())
                    comp = converter_valor(campos["comp"].get())
                    kg_metro = ((l1 * l2) / 1000) * material - (((l1 - esp * 2) * (l2 - esp * 2)) / 1000) * material
                    kg_total = (kg_metro * comp) / 1000
                elif calculo == "CT":
                    l1 = converter_valor(campos["l1"].get())
                    l2 = converter_valor(campos["l2"].get())
                    esp = converter_valor(campos["esp"].get())
                    comp = converter_valor(campos["comp"].get())
                    kg_metro = (((l1 * esp) + ((l2 - esp) * esp)) * material * 1000) / 1000000000
                    kg_total = (kg_metro * comp) / 1000
                elif calculo == "BR":
                    di = converter_valor(campos["di"].get())
                    comp = converter_valor(campos["comp"].get())
                    kg_metro = ((di / 2000)**2) * 3.1416 * material * 1000
                    kg_total = (kg_metro * comp) / 1000
                elif calculo == "BRQ":
                    l1 = converter_valor(campos["l1"].get())
                    l2 = converter_valor(campos["l2"].get())
                    comp = converter_valor(campos["comp"].get())
                    kg_metro = (l1 / 1000) * (l2 / 1000) * comp * material
                    kg_total = (kg_metro * comp) / 1000
                elif calculo == "BS":
                    fa = converter_valor(campos["fa"].get())
                    comp = converter_valor(campos["comp"].get())
                    kg_metro = (fa / 1000) * 0.57735 * (fa / 2000) * 6 * comp * material
                    kg_total = (kg_metro * comp) / 1000
                else:
                    raise ValueError("Selecione um tipo de cálculo válido.")
                
                resultado_label.config(text=f"Peso total: {kg_total:.3f} Kg")
            except ValueError as ve:
                resultado_label.config(text=f"Erro: {ve}")
            except Exception as e:
                resultado_label.config(text=f"Erro inesperado: {e}")

        # Interface gráfica para Tabela de Pesos
        tk.Label(janela, text="Tabela de pesos").pack(pady=5)
        tipo_combobox = ttk.Combobox(janela, values=list(tipo_calculo.keys()), width=30)
        tipo_combobox.pack()
        tipo_combobox.bind("<<ComboboxSelected>>", atualizar_campos)

        tk.Label(janela, text="Material").pack(pady=5)
        material_combobox = ttk.Combobox(janela, values=list(material_aplicacao.keys()), width=30)
        material_combobox.pack()

        campos_frame = tk.Frame(janela)
        campos_frame.pack(pady=10)

        tk.Button(janela, text="Calcular", command=calcular, width=30).pack(pady=10)
        resultado_label = tk.Label(janela, text="", justify="left", bg="white", width=40, height=5, relief="sunken", anchor="nw", padx=5, pady=5)
        resultado_label.pack(pady=10)
        campos = {}

    # Função para criar a janela de Mangueiras
    def criar_janela_mangueira(janela):
        def calcular_mangueira():
            try:
                # Obtendo valores do formulário
                diametro = diametro_var.get()
                comprimento = float(entry_comprimento.get())
                material_arame = material_arame_var.get().lower()
                tipo_mangueira = int(tipo_mangueira_var.get())
                material_mangueira = material_mangueira_var.get()

                diametros = {
                    "1/4": 6.33,
                    "3/8": 9.525,
                    "1/2": 12.7,
                    "3/4": 19.05,
                    "1": 25.4,
                    "1.1/2": 38.1,
                    "2": 50.8,
                    "2.1/2": 63.5,
                    "3": 76.2,
                    "4": 101.6,
                    "6": 152.4,
                    "8": 203.2,
                    }
                materiais = {
                    "aco galvanizado": 7.2,
                    "aco revestido": 8.8,
                    "aco inox 304": 8.4,
                    "aco inox 316": 8.1,
                    "aco inox 304 revestido": 7,
                    }
                diametro_interno = diametros.get(diametro)
                if diametro_interno is None:
                    messagebox.showerror("Erro", "Diâmetro inválido. Escolha um válido.")
                    return

                mat_arame = materiais.get(material_arame)
                if mat_arame is None:
                    messagebox.showerror("Erro", "Material de arame inválido. Escolha um válido.")
                    return

                # Cálculo inicial
                esp_ate_4, esp_ate_6, esp_ate_8 = 15, 22, 28

                if diametro_interno <= 25.4:  # Menor que 1 pol
                    diametro_externo = diametro_interno + esp_ate_4
                    arame_interno = 2
                    arame_externo = 2
                    passo_arame = 10
                    di_corda = '1/4"'
                    di_corda_value = 6.35
                    gm_corda = 20
                    pead_porcent = 1
                elif diametro_interno == 38.1 or diametro_interno == 50.8:  # 1.1/2 e 2 pol
                    diametro_externo = diametro_interno + esp_ate_4
                    arame_interno = 2.5
                    arame_externo = 2.5
                    passo_arame = 11
                    di_corda = '1/4"'
                    di_corda_value = 6.35
                    gm_corda = 20
                    pead_porcent = 1
                elif diametro_interno == 63.5:  # 2.1/2 pol
                    diametro_externo = diametro_interno + esp_ate_4
                    arame_interno = 3
                    arame_externo = 3
                    passo_arame = 14
                    di_corda = '1/4"'
                    di_corda_value = 6.35
                    gm_corda = 20
                    pead_porcent = 1
                elif diametro_interno == 76.2:  # 3 pol
                    diametro_externo = diametro_interno + esp_ate_4
                    arame_interno = 3.4
                    arame_externo = 3.4
                    passo_arame = 14
                    di_corda = '1/4"'
                    di_corda_value = 6.35
                    gm_corda = 20
                    pead_porcent = 1
                elif diametro_interno == 101.6: # 4 pol
                    diametro_externo = diametro_interno + esp_ate_4
                    arame_interno = 3.4
                    arame_externo = 3.4
                    passo_arame = 14
                    di_corda = '3/8"'
                    di_corda_value = 9.525
                    gm_corda = 48
                    pead_porcent = 1
                elif diametro_interno == 152.4: # 6 pol
                    diametro_externo = diametro_interno + esp_ate_6
                    arame_interno = 5
                    arame_externo = 6
                    passo_arame = 18
                    di_corda = '1/2"'
                    di_corda_value = 12.7
                    gm_corda = 77
                    pead_porcent = 1.5
                elif diametro_interno == 203.2: # 8 pol
                    diametro_externo = diametro_interno + esp_ate_8
                    arame_interno = 6
                    arame_externo = 6
                    passo_arame = 20
                    di_corda = '1/2"'
                    di_corda_value = 12.7
                    gm_corda = 77
                    pead_porcent = 1.5
                if tipo_mangueira == 1:
                    if material_mangueira == "PP":
                        if diametro_interno <= 25.4:  # 1 pol
                            camadas = 4
                            rafia = 3
                            pelicula_ptfe = 0
                            larg_rafia = 150
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 38.1 or  diametro_interno == 50.8:  # 1.1/2 e 2 pol
                            camadas = 6
                            rafia = 3
                            pelicula_ptfe = 0
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 63.5 or diametro_interno == 76.2: #  2.1/2 3 pol
                            camadas = 6
                            rafia = 3
                            pelicula_ptfe = 0
                            larg_rafia = 500
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 101.6: # 4 pol
                            camadas = 12
                            rafia = 5
                            pelicula_ptfe = 0
                            larg_rafia = 500
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 152.4: # 6 pol
                            camadas = 24
                            rafia = 10
                            pelicula_ptfe = 0
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 0
                        elif diametro_interno == 203.2: # 8 pol
                            camadas = 28
                            rafia = 10
                            pelicula_ptfe = 0
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 0
                if tipo_mangueira == 2:
                    if material_mangueira == "PP":
                        if diametro_interno <= 25.4:  # 1 pol
                            camadas = 5
                            rafia = 3
                            pelicula_ptfe = 0
                            larg_rafia = 150
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 38.1 or  diametro_interno == 50.8:  # 1.1/2 e 2 pol
                            camadas = 6
                            rafia = 4
                            pelicula_ptfe = 0
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 63.5 or diametro_interno == 76.2: #  2.1/2 3 pol
                            camadas = 10
                            rafia = 3
                            pelicula_ptfe = 0
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 101.6: # 4 pol
                            camadas = 12
                            rafia = 5
                            pelicula_ptfe = 0
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 152.4: # 6 pol
                            camadas = 24
                            rafia = 10
                            pelicula_ptfe = 0
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 0
                        elif diametro_interno == 203.2: # 8 pol
                            camadas = 28
                            rafia = 14
                            pelicula_ptfe = 0
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 0
                if tipo_mangueira == 3:
                    if material_mangueira == "PP":
                        if diametro_interno <= 25.4:  # 1 pol
                            camadas = 5
                            rafia = 3
                            pelicula_ptfe = 0
                            larg_rafia = 150
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 38.1 or  diametro_interno == 50.8:  # 1.1/2 e 2 pol
                            camadas = 6
                            rafia = 3
                            pelicula_ptfe = 0
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 63.5 or diametro_interno == 76.2: #  2.1/2 3 pol
                            camadas = 10
                            rafia = 5
                            pelicula_ptfe = 0
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 101.6: # 4 pol
                            camadas = 12
                            rafia = 7
                            pelicula_ptfe = 0
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 0
                        elif diametro_interno == 152.4: # 6 pol
                            camadas = 24
                            rafia = 10
                            pelicula_ptfe = 0
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 0
                        elif diametro_interno == 203.2: # 8 pol
                            camadas = 28
                            rafia = 18
                            pelicula_ptfe = 0
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 0
                if tipo_mangueira == 1:
                    if material_mangueira == "PTFE":
                        if diametro_interno <= 25.4:  # 1 pol
                            camadas = 4
                            rafia = 3
                            pelicula_ptfe = 1
                            larg_rafia = 150
                            gm2_rafia = 220
                            larg_teflon = 150
                        elif diametro_interno == 38.1 or  diametro_interno == 50.8:  # 1.1/2 e 2 pol
                            camadas = 6
                            rafia = 3
                            pelicula_ptfe = 1
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 250
                        elif diametro_interno == 63.5 or diametro_interno == 76.2: #  2.1/2 3 pol
                            camadas = 6
                            rafia = 3
                            pelicula_ptfe = 1
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 400
                        elif diametro_interno == 101.6: # 4 pol
                            camadas = 12
                            rafia = 5
                            pelicula_ptfe = 1
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 400
                        elif diametro_interno == 152.4: # 6 pol
                            camadas = 24
                            rafia = 10
                            pelicula_ptfe = 1
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 800
                        elif diametro_interno == 203.2: # 8 pol
                            camadas = 28
                            rafia = 10
                            pelicula_ptfe = 1
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 800
                if tipo_mangueira == 2:
                    if material_mangueira == "PTFE":
                        if diametro_interno <= 25.4:  # 1 pol
                            camadas = 5
                            rafia = 3
                            pelicula_ptfe = 1
                            larg_rafia = 150
                            gm2_rafia = 160
                            larg_teflon = 150
                        elif diametro_interno == 38.1 or  diametro_interno == 50.8:  # 1.1/2 e 2 pol
                            camadas = 6
                            rafia = 4
                            pelicula_ptfe = 1
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 250
                        elif diametro_interno == 63.5 or diametro_interno == 76.2: #  2.1/2 3 pol
                            camadas = 10
                            rafia = 3
                            pelicula_ptfe = 1
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 400
                        elif diametro_interno == 101.6: # 4 pol
                            camadas = 12
                            rafia = 5
                            pelicula_ptfe = 1
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 400
                        elif diametro_interno == 152.4: # 6 pol
                            camadas = 24
                            rafia = 10
                            pelicula_ptfe = 1
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 800
                        elif diametro_interno == 203.2: # 8 pol
                            camadas = 28
                            rafia = 14
                            pelicula_ptfe = 1
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 800
                if tipo_mangueira == 3:
                    if material_mangueira == "PTFE":
                        if diametro_interno <= 25.4:  # 1 pol
                            camadas = 5
                            rafia = 3
                            pelicula_ptfe = 1
                            larg_rafia = 150
                            gm2_rafia = 160
                            larg_teflon = 150
                        elif diametro_interno == 38.1 or  diametro_interno == 50.8:  # 1.1/2 e 2 pol
                            camadas = 6
                            rafia = 3
                            pelicula_ptfe = 1
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 250
                        elif diametro_interno == 63.5 or diametro_interno == 76.2: #  2.1/2 3 pol
                            camadas = 10
                            rafia = 5
                            pelicula_ptfe = 1
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 400
                        elif diametro_interno == 101.6: # 4 pol
                            camadas = 12
                            rafia = 7
                            pelicula_ptfe = 1
                            larg_rafia = 400
                            gm2_rafia = 160
                            larg_teflon = 400
                        elif diametro_interno == 152.4: # 6 pol
                            camadas = 24
                            rafia = 10
                            pelicula_ptfe = 1
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 800
                        elif diametro_interno == 203.2: # 8 pol
                            camadas = 28
                            rafia = 18
                            pelicula_ptfe = 1
                            larg_rafia = 800
                            gm2_rafia = 220
                            larg_teflon = 800
                if tipo_mangueira == 4:
                    if material_mangueira == "Nylon":
                        if diametro_interno <= 25.4:  # 1 pol
                            camadas = 5
                            rafia = 0
                            pelicula_ptfe = 2
                            larg_rafia = 0
                            gm2_rafia = 0
                            larg_teflon = 150
                        elif diametro_interno == 38.1 or  diametro_interno == 50.8:  # 1.1/2 e 2 pol
                            camadas = 5
                            rafia = 0
                            pelicula_ptfe = 2
                            larg_rafia = 0
                            gm2_rafia = 0
                            larg_teflon = 250
                        elif diametro_interno == 63.5 or diametro_interno == 76.2: #  2.1/2 3 pol
                            camadas = 6
                            rafia = 0
                            pelicula_ptfe = 2
                            larg_rafia = 0
                            gm2_rafia = 0
                            larg_teflon = 400
                        elif diametro_interno == 101.6: # 4 pol
                            camadas = 7
                            rafia = 0
                            pelicula_ptfe = 2
                            larg_rafia = 0
                            gm2_rafia = 0
                            larg_teflon = 400
                        elif diametro_interno == 152.4: # 6 pol
                            camadas = 8
                            rafia = 0
                            pelicula_ptfe = 2
                            larg_rafia = 0
                            gm2_rafia = 0
                            larg_teflon = 800
                        elif diametro_interno == 203.2: # 8 pol
                            camadas = 9
                            rafia = 0
                            pelicula_ptfe = 2
                            larg_rafia = 0
                            gm2_rafia = 0
                            larg_teflon = 800
                gab_mm = passo_arame + arame_interno
                raio_int_mm = (diametro_interno/2) + arame_interno/2
                comp_int_mm = (1.1*comprimento*math.sqrt(abs(((2*3.14*raio_int_mm) ** 2) + (gab_mm**2)))/gab_mm)/1000
                cm3_int_mm = ((3.14*((arame_interno/10)**2))/4) * (comp_int_mm*100)
                peso_arame_int = (mat_arame*cm3_int_mm)/1000
                raio_ext_mm = (diametro_externo/2) - arame_externo/2
                comp_ext_mm = (math.sqrt(abs(((2*3.14*raio_ext_mm)**2)+(gab_mm**2)))*comprimento/gab_mm)/1000
                cm3_ext_mm = ((3.14*((arame_externo/10)**2))/4) * (comp_ext_mm*100)
                peso_arame_ext = ((mat_arame*cm3_ext_mm)/1000)*1.1
                comp_filme = (math.sqrt(abs(((2*3.14*diametro_externo)**2)+((800*1.8)**2)))*comprimento/(800*1.8))/1000
                area_filme = comp_filme*0.8
                peso_filme = ((area_filme*46)/1000)*camadas*1.2
                try:
                    comp_rafia = (math.sqrt(abs(((2*3.14*((diametro_interno+diametro_externo)/2))**2)+((larg_rafia*1.8)**2)))*comprimento/(larg_rafia*1.8))/1000
                    area_rafia = comp_rafia*(larg_rafia/1000)
                    peso_rafia = ((area_rafia*gm2_rafia)/1000*rafia)*1.2
                except ZeroDivisionError:
                    comp_rafia = 0
                    area_rafia = 0
                    peso_rafia = 0
                try:
                    comp_teflon = (math.sqrt(abs(((2*3.14*diametro_externo)**2)+((larg_teflon*1.8)**2)))*comprimento/(larg_teflon*1.8))/1000
                    area_teflon = comp_teflon*(larg_teflon/1000)
                    peso_teflon = ((area_teflon*235)/1000)*pelicula_ptfe*1.2
                except ZeroDivisionError:
                    comp_teflon = 0
                    area_teflon = 0
                    peso_teflon = 0
                comp_corda = (math.sqrt(abs(((2*3.14*raio_ext_mm)**2)+di_corda_value**2))*comprimento/di_corda_value)/1000
                corda = (comp_corda*gm_corda)/1000
                pead = pead_porcent*comprimento

                resultado_box.delete("1.0", tk.END)  # Limpa o conteúdo anterior
                resultado = f' - Diamentro interno: {diametro_interno}\n - Diâmetro externo: {diametro_externo}\n\n - Arame interno : {round(peso_arame_int,2)} Kg\n - Arame externo : {round(peso_arame_ext,2)} Kg\n\n - Comprimento : {int(comprimento)} mm\n - Diametro do arame : {passo_arame}\n\n - QTDE. Camadas filme : {camadas}\n - Filme : {round(peso_filme,2)} Kg\n\n - QTDE. Camadas rafia : {rafia}\n - Rafia : {round(peso_rafia,2)} Kg\n\n - QTDE. Pelicula teflon : {pelicula_ptfe}\n - Pelicula teflon : {round(peso_teflon)} Kg\n\n - QTDE. PEAD : {int(pead)} mm\n\n - Corda de : {di_corda}\n - QTDE. Corda: {round(corda,2)} Kg'
                resultado_box.insert(tk.END, resultado)

            except ValueError:
                resultado_box.delete("1.0", tk.END)

        # Entrada de Dados
        tk.Label(janela, text="Diâmetro:").grid(row=0, column=0, pady=5, sticky="e")
        global diametro_var, entry_comprimento, material_arame_var, tipo_mangueira_var, material_mangueira_var, resultado_box
        diametro_var = tk.StringVar()
        diametro_menu = ttk.Combobox(janela, textvariable=diametro_var, values=[
            "1/4", "3/8", "1/2", "3/4", "1", "1.1/2", "2", "2.1/2", "3", "4", "6", "8"
        ])
        diametro_menu.grid(row=0, column=1, pady=5)

        tk.Label(janela, text="Comprimento (mm):").grid(row=1, column=0, pady=5, sticky="e")
        entry_comprimento = tk.Entry(janela)
        entry_comprimento.grid(row=1, column=1, pady=5)

        tk.Label(janela, text="Material do Arame:").grid(row=2, column=0, pady=5, sticky="e")
        material_arame_var = tk.StringVar()
        material_arame_menu = ttk.Combobox(janela, textvariable=material_arame_var, values=[
            "aco galvanizado", "aco revestido", "aco inox 304", "aco inox 316", "aco inox 304 revestido"
        ])
        material_arame_menu.grid(row=2, column=1, pady=5)

        tk.Label(janela, text="Tipo de Mangueira:").grid(row=3, column=0, pady=5, sticky="e")
        tipo_mangueira_var = tk.StringVar()
        tipo_mangueira_menu = ttk.Combobox(janela, textvariable=tipo_mangueira_var, values=["1", "2", "3"])
        tipo_mangueira_menu.grid(row=3, column=1, pady=5)

        tk.Label(janela, text="Material da Mangueira:").grid(row=4, column=0, pady=5, sticky="e")
        material_mangueira_var = tk.StringVar()
        material_mangueira_menu = ttk.Combobox(janela, textvariable=material_mangueira_var, values=["PP", "PTFE"])
        material_mangueira_menu.grid(row=4, column=1, pady=5)

        # Botão de Cálculo
        btn_calcular = tk.Button(janela, text="Calcular", command=calcular_mangueira)
        btn_calcular.grid(row=5, columnspan=2, pady=20)
        
        # Caixa de Resultado
        tk.Label(janela, text="Resultado:").grid(row=6, column=0, columnspan=2)
        resultado_box = tk.Text(janela, height=22, width=60)
        resultado_box.grid(row=7, column=0, columnspan=2, pady=10)

    # Tela principal
    app = tk.Tk()
    app.title("Menu Principal")
    app.geometry("400x400")

    # Botões da tela inicial
    tk.Label(app, text="Selecione o calculo desejado:", font=("Arial", 14)).pack(pady=10)

    tk.Button(app, text="Calcular Pesos", command=abrir_tabela_pesos, width=20, bg="green", fg="white").pack(pady=10)
    # tk.Button(app, text="Calcular Mangueiras", command=abrir_mangueiras, width=20, bg="blue", fg="white").pack(pady=10)

    app.mainloop()

def importados_geral():
    # Configuração do banco de dados
    db = create_engine('sqlite:///Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\banco de dados\\Importados.db')
    Session = sessionmaker(bind=db)
    db_session = Session()
    Base1 = declarative_base()

    # Definição do modelo do banco de dados
    class Produtos(Base1):
        __tablename__ = 'Produtos'
        Id = Column('Id', Integer, primary_key=True, autoincrement=True)
        Empresa = Column('Empresa', String)
        Produto = Column('Produto', String)
        Descricao = Column('Descrição', String)
        preco = Column('preço', Float)

    # Função para obter cotações
    def obter_cotacoes():
        cotacoes = requests.get('https://economia.awesomeapi.com.br/last/USD-BRL,EUR-BRL,BTC-BRL')
        cotacoes = cotacoes.json()
        return float(cotacoes['USDBRL']['bid']), float(cotacoes['EURBRL']['bid'])

    cotacao_dolar, cotacao_euro = obter_cotacoes()

    # Função para buscar produtos no banco de dados
    def buscar_produtos():
        nome_produto = entrada_nome.get().strip()  # Remover espaços em branco extras
        if not nome_produto:
            messagebox.showwarning("Aviso", "Por favor, insira ao menos uma parte do nome do produto.")
            return

        filtro = filtro_combobox.get()
        nome_final_produto = f'%{nome_produto}%'
        produtos_encontrados = []

        if filtro == 'Código':
            produtos_encontrados = db_session.query(Produtos).filter(Produtos.Produto.like(nome_final_produto)).all()
        elif filtro == 'Empresa':
            produtos_encontrados = db_session.query(Produtos).filter(Produtos.Empresa.like(nome_final_produto)).all()
        else:  # Filtro por Descrição
            produtos_encontrados = db_session.query(Produtos).filter(Produtos.Descricao.like(nome_final_produto)).all()

        if not produtos_encontrados:
            messagebox.showerror("Erro", f"Nenhum produto encontrado com o termo '{nome_produto}'.")
            return

        # Preenchendo os resultados na tabela
        for row in resultados_tree.get_children():
            resultados_tree.delete(row)

        for produto in produtos_encontrados:
            preco_final = calcular_preco(produto)
            resultados_tree.insert('', tk.END, values=(
                produto.Id, produto.Empresa, produto.Produto, produto.Descricao, f"R$ {preco_final:.2f}"
            ))

    def calcular_preco(produto):
        if produto.Empresa == 'OPW':
            return round(produto.preco * cotacao_dolar * 1.10 * 1.65 * 0.534, 2)
        return round(produto.preco * cotacao_dolar * 1.10 * 1.65, 2)

    # Interface gráfica com Tkinter
    app = tk.Tk()
    app.title("Consulta de Produtos Importados")

    # Layout da aplicação
    frame_filtro = tk.Frame(app)
    frame_filtro.pack(pady=10)

    tk.Label(frame_filtro, text="Nome do Produto:").grid(row=0, column=0, padx=5, pady=5)
    entrada_nome = tk.Entry(frame_filtro, width=30)
    entrada_nome.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(frame_filtro, text="Filtro:").grid(row=0, column=2, padx=5, pady=5)
    filtro_combobox = ttk.Combobox(frame_filtro, values=["Código", "Empresa", "Descrição"])
    filtro_combobox.grid(row=0, column=3, padx=5, pady=5)
    filtro_combobox.set("Código")

    tk.Button(frame_filtro, text="Buscar", command=buscar_produtos).grid(row=0, column=4, padx=5, pady=5)

    # Tabela para exibir resultados
    resultados_tree = ttk.Treeview(app, columns=("ID", "Empresa", "Produto", "Descrição", "Preço"), show='headings')
    resultados_tree.heading("ID", text="ID")
    resultados_tree.heading("Empresa", text="Empresa")
    resultados_tree.heading("Produto", text="Produto")
    resultados_tree.heading("Descrição", text="Descrição")
    resultados_tree.heading("Preço", text="Preço")

    resultados_tree.column("ID", width=50, anchor=tk.CENTER)
    resultados_tree.column("Empresa", width=100, anchor=tk.W)
    resultados_tree.column("Produto", width=150, anchor=tk.W)
    resultados_tree.column("Descrição", width=250, anchor=tk.W)
    resultados_tree.column("Preço", width=100, anchor=tk.E)

    resultados_tree.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    # Rodar o aplicativo
    app.mainloop()

def setores_de_estoque_geral():
    url = 'Z:\\ISO 9000 - SGQ\\12-SISTEMA\\Sistema\\planilhas\\Setores de estoque.xlsx'
    df = pd.read_excel(url)

    def estoque_uso_e_consumo():
        filtro = df['Nome do setor de estoque'] == 'ESTOQUE USO E CONSUMO'
        df_filtrado = df[filtro]
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        output_file = 'Estoque uso e consumo.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def estoque_processo():
        filtro = df['Nome do setor de estoque'] == 'ESTOQUE PROCESSO'
        df_filtrado = df[filtro]
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        output_file = 'Estoque processo.xlsx'
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def estoque_em_terceiros():
        filtro = df['Nome do setor de estoque'] == 'ESTOQUE EM TERCEIROS'
        df_filtrado = df[filtro]
        output_file = 'Estoque em terceiros.xlsx'
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def estoque_retono_terceiros():
        filtro = df['Nome do setor de estoque'] == 'ESTOQUE RETORNO TERCEIROS'
        df_filtrado = df[filtro]
        output_file = 'Estoque retorno terceiros.xlsx'
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def estoque_de_clientes():
        filtro = df['Nome do setor de estoque'] == 'ESTOQUE DE CLIENTES'
        df_filtrado = df[filtro]
        output_file = 'Estoque de clientes.xlsx'
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def estoque_almoxarifado_central():
        filtro = df['Nome do setor de estoque'] == 'ESTOQUE ALMOXARIFADO CENTRAL'
        df_filtrado = df[filtro]
        output_file = 'Estoque almoxarifado central.xlsx'
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def estoque_expedicao_acabados():
        filtro = df['Nome do setor de estoque'] == 'ESTOQUE EXPEDIÇÃO ACABADOS'
        df_filtrado = df[filtro]
        output_file = 'Estoque expedição acabados.xlsx'
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def ativos_imobilizados():
        filtro = df['Nome do setor de estoque'] == 'ATIVOS IMOBILIZADOS'
        df_filtrado = df[filtro]
        output_file = 'Ativos imobilizados.xlsx'
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def almoxarifado_PHD():
        filtro = df['Nome do setor de estoque'] == 'ALMOXARIFADO PHD'
        df_filtrado = df[filtro]
        output_file = 'Almoxarifado PHD.xlsx'
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def estoque_processo_PHD():
        filtro = df['Nome do setor de estoque'] == 'EST. PROCESSO PHD'
        df_filtrado = df[filtro]
        output_file = 'Estoque processo PHD.xlsx'
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def acabados_PHD():
        filtro = df['Nome do setor de estoque'] == 'ACABADOS PHD'
        df_filtrado = df[filtro]
        output_file = 'Acabados PHD.xlsx'
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def estoque_almoxarifado_pre_producao():
        filtro = df['Nome do setor de estoque'] == 'ESTOQUE ALMOXARIFADO PRÉ-PRODUÇÃO'
        df_filtrado = df[filtro]
        output_file = 'Estoque almoxarifado pré-produção.xlsx'
        total_custo = sum(df_filtrado['Custo médio total'])
        df_filtrado['80 - 20 (%)'] = (df_filtrado['Custo médio total'] / total_custo)
        df_filtrado = df_filtrado.sort_values(by='Custo médio total', ascending=False)
        pasta_destino = r'Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques'
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_completo = os.path.join(pasta_destino, output_file)
        df_filtrado.to_excel(caminho_completo, index=False)
        wb = load_workbook(caminho_completo)
        ws = wb.active
        tabela = Table(displayName="Tabela", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col[0].value == "Custo médio unitário" or col[0].value == "Custo médio total":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '#,##0.00' 
            if col[0].value == "80 - 20 (%)":
                for cell in col:
                    if cell.row != 1: 
                        cell.number_format = '0.00%' 
            if col[0].value == "Descrição do produto":
                ws.column_dimensions[col_letter].width = 70
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2 
        wb.save(caminho_completo)

    def instrucao():
        pyautogui.alert("planilhas são salvas em Z:\ISO 9000 - SGQ\9 - PPCP\Estoque\Estoques")

    def criar_interface():
        janela = tk.Tk()
        janela.title("Estoques")
        janela.geometry("400x550")
        
        tk.Label(janela, text="Selecione o estoque desejado:", font=("Arial", 14)).pack(pady=10)

        tk.Button(janela, text="Instrução", command=instrucao, width=30, bg="purple", fg="white").pack(pady=5)
        tk.Button(janela, text="Estoque almoxarifado central", command=estoque_almoxarifado_central, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="Estoque almoxarifado pré-produção", command=estoque_almoxarifado_pre_producao, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="Estoque expedição acabados", command=estoque_expedicao_acabados, width=30, bg="green", fg="white").pack(pady=5)
        tk.Button(janela, text="Estoque Processo", command=estoque_processo, width=30, bg="yellow", fg="black").pack(pady=5)
        tk.Button(janela, text="Ativo imobilizado", command=ativos_imobilizados, width=30, bg="yellow", fg="black").pack(pady=5)
        tk.Button(janela, text="Estoque uso e consumo", command=estoque_uso_e_consumo, width=30, bg="yellow", fg="black").pack(pady=5)
        tk.Button(janela, text="Estoque em terceiros", command=estoque_em_terceiros, width=30, bg="blue", fg="white").pack(pady=5)
        tk.Button(janela, text="Estoque retorno terceiros", command=estoque_retono_terceiros, width=30, bg="blue", fg="white").pack(pady=5)
        tk.Button(janela, text="Estoque de clientes", command=estoque_de_clientes, width=30, bg="blue", fg="white").pack(pady=5)
        tk.Button(janela, text="Almoxarifado PHD", command=almoxarifado_PHD, width=30, bg="red", fg="white").pack(pady=5)
        tk.Button(janela, text="Estoque processo PHD", command=estoque_processo_PHD, width=30, bg="red", fg="white").pack(pady=5)
        tk.Button(janela, text="Acabados_PHD", command=acabados_PHD, width=30, bg="red", fg="white").pack(pady=5)

        janela.mainloop()

    criar_interface()

def extracao_lm_excel():
    pasta_saida = r'Z:\PUBLICO\Araujo\web_itens_lm'

    def entrar_lm_e_enviar_para_o_excel():
        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico) 
        usuario = entry_usuario.get() 
        senha = entry_senha.get()
        cod = entry_codigo.get()
        data_atual = datetime.now().strftime("%Y-%m-%d")

        def login(usuario, senha):
            tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
            tela_usuario.send_keys(usuario)
            tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
            tela_senha.send_keys(senha)    
            tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_produtos():
            navegador.get('https://tspro.nomus.com.br/tspro/Produto.do?metodo=Pesquisar') # Produtos

        def entra_produto(cod):
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'divMinimizavel')]/table/tbody/tr[2]/td[1]/input"))) # Código do produto
            produto_01.send_keys(cod)

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisar"]'))) # Buscar produtos
            botao_buscar.click()

            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{cod}']"))) #Seleciona o elemento
            botao_buscar.click()

            xpaths = ['//*[@id="produtoAtivoAguardandoLiberacao_itemSubMenu_acessarListaMateriais"]', '//*[@id="produtoAtivoLiberado_itemSubMenu_acessarListaMateriais"]']
            for xpath in xpaths:
                try:
                    botao_entrar_lm = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    botao_entrar_lm.click()
                    break
                except Exception:
                    pass    

        def abrir_lm():
            botao_lm = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_Funcoes_especiais"]'))) # Buscar produtos
            botao_lm.click()

            time.sleep(1)

            botao_lm2 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_carregar_todos_niveis_lista"]'))) # Buscar produtos
            botao_lm2.click()

            time.sleep(1)

            # Espera até que a tabela seja carregada
            table_xpath = '/html/body/div[2]/div[4]/div/form/div[2]/table'
            table = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.XPATH, table_xpath)))

            # Extrai os dados da tabela
            rows = table.find_elements(By.TAG_NAME, 'tr')
            data = []
            for row in rows:
                cols = row.find_elements(By.TAG_NAME, 'td')
                row_data = [col.text for col in cols]
                data.append(row_data)

            # Cria um DataFrame com os dados
            df = pd.DataFrame(data)
            df.drop(index=df.index[:130], inplace=True)
            df.drop(df.columns[[0, 1]], axis=1, inplace=True)
            df.dropna(inplace=False)

            # Exporta os dados para um arquivo Excel
            df.to_excel(os.path.join(pasta_saida, f'{cod} - {data_atual}.xlsx'), index=False)

            # Fecha o navegador
            navegador.quit()

        def planilha():
            caminho_arquivo = r"Z:\ISO 9000 - SGQ\9 - PPCP\.PCP\Extrair LM\Extração.xlsx"
            excel = win32com.client.Dispatch("Excel.Application")
            workbook = excel.Workbooks.Open(caminho_arquivo)
            workbook.RefreshAll()
            time.sleep(1)
            nome_da_aba = "Consulta PAI"
            try:
                sheet = workbook.Worksheets(nome_da_aba)
            except Exception as e:
                print(f"Erro: A aba '{nome_da_aba}' não foi encontrada.")
                excel.Quit()
                exit()
            sheet.Range("F2").Value = cod
            excel.Visible = True
        entrar_nomus()
        tela_produtos()
        entra_produto(cod)
        abrir_lm()
        planilha()

    app = tk.Tk()
    app.title("Engenharia")
    app.geometry("400x300")
    app.config(bg="lightblue")  

    tk.Label(app, text="Selecione a opção desejada: ", font=("Arial", 14), bg="lightblue").pack(pady=5)
    label_codigo = tk.Label(app, text="Código Pai:", bg="lightblue")
    label_codigo.pack(padx=5, pady=5)

    entry_codigo = tk.Entry(app, bg="white", bd=2, relief="solid", justify="center")
    entry_codigo.pack(padx=5, pady=5)

    label_usuario = tk.Label(app, text="Usuário:", bg="lightblue")
    label_usuario.pack(padx=5, pady=5)

    entry_usuario = tk.Entry(app, bg="white", bd=2, relief="solid", justify="center")
    entry_usuario.pack(padx=5, pady=5)

    label_senha = tk.Label(app, text="Senha:",  bg="lightblue")
    label_senha.pack(padx=5, pady=5)

    entry_senha = tk.Entry(app, show="*", bg="white", bd=2, relief="solid", justify="center")
    entry_senha.pack(padx=5, pady=5)

    tk.Button(app, text="Extrair Lista", command=entrar_lm_e_enviar_para_o_excel, width=20, bg="gold", fg="black").pack(pady=10)

    app.mainloop()

def baixar_desenhos():
    def entrar_lm_e_enviar_para_o_excel():
        servico = Service(ChromeDriverManager().install())
        navegador = webdriver.Chrome(service=servico) 
        usuario = entry_usuario.get() 
        senha = entry_senha.get()
        op = entry_op.get()

        def login(usuario, senha):
            tela_usuario = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="campologin"]')))
            tela_usuario.send_keys(usuario)
            tela_senha = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login_form"]/div/main/div/section[3]/div[2]/input')))
            tela_senha.send_keys(senha)    
            tela_senha.send_keys(Keys.RETURN)

        def entrar_nomus():
            navegador.get('https://tspro.nomus.com.br/tspro/Login.do?metodo=PreLogin') # Tela inicial
            login(usuario, senha)

        def tela_ordens():
            navegador.get('https://tspro.nomus.com.br/tspro/Ordem.do?metodo=pesquisarPaginado') # Produtos

        def entra_ordem(op):
            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_exibir_todos"]'))) # Buscar produtos
            botao_buscar.click()
            produto_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[4]/div/form/div[6]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td[1]/input"))) # Código do produto
            produto_01.send_keys(op)
            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_pesquisarpaginado"]'))) # Buscar produtos
            botao_buscar.click()
            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, f"//*[text()='{op}']"))) #Seleciona o elemento
            botao_buscar.click()
            botao_buscar_01 = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, '_itemSubMenu_gerarZipArquivosAnexosOrdem')]"))) # Buscar produtos
            botao_buscar_01.click()
            time.sleep(5)
            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="marcaredesmarcar"]'))) # Buscar produtos
            botao_buscar.click()
            time.sleep(0.5)
            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_Acoes"]'))) # Buscar produtos
            botao_buscar.click()
            time.sleep(0.5)
            botao_buscar = WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="botao_botao.gerarpdfordensemlote"]'))) # Buscar produtos
            botao_buscar.click()
            time.sleep(5)
            navegador.quit()

        def winrar():
            def extrair_arquivos(arquivo_rar, pasta_destino):
                try:
                    temp_pasta = os.path.join(pasta_destino, "temp")
                    os.makedirs(temp_pasta, exist_ok=True)
                    patoolib.extract_archive(arquivo_rar, outdir=temp_pasta)
                    for root, _, files in os.walk(temp_pasta):
                        for file in files:
                            origem = os.path.join(root, file)
                            destino = os.path.join(pasta_destino, file)
                            if os.path.exists(destino):
                                nome, extensao = os.path.splitext(file)
                                contador = 1
                                novo_nome = f"{nome}_{contador}{extensao}"
                                while os.path.exists(os.path.join(pasta_destino, novo_nome)):
                                    contador += 1
                                    novo_nome = f"{nome}_{contador}{extensao}"
                                destino = os.path.join(pasta_destino, novo_nome)
                            shutil.move(origem, destino)
                    shutil.rmtree(temp_pasta)
                    print(f"Arquivos extraídos de {arquivo_rar} para {pasta_destino}.")
                except Exception as e:
                    print(f"Erro ao extrair {arquivo_rar}: {e}")
            pasta_downloads = os.path.expanduser("~/Downloads")
            pasta_destino = os.path.join(pasta_downloads, "Extraidos")
            os.makedirs(pasta_destino, exist_ok=True)
            for arquivo in os.listdir(pasta_downloads):
                if arquivo.endswith(".rar") or arquivo.endswith(".zip"):
                    caminho_arquivo = os.path.join(pasta_downloads, arquivo)
                    extrair_arquivos(caminho_arquivo, pasta_destino)
            for arquivo in os.listdir(pasta_downloads):
                caminho_arquivo = os.path.join(pasta_downloads, arquivo)
                try:
                    if os.path.isfile(caminho_arquivo):
                        os.remove(caminho_arquivo)
                        print(f"Deletado: {caminho_arquivo}")
                except Exception as e:
                    print(f"Erro ao deletar {caminho_arquivo}: {e}")

        entrar_nomus()
        tela_ordens()
        entra_ordem(op)
        winrar()

    app = tk.Tk()
    app.title("Engenharia")
    app.geometry("400x350")
    app.config(bg="lightblue")  

    tk.Label(app, text="Selecione a opção desejada: ", font=("Arial", 14), bg="lightblue").pack(pady=5)
    label_op = tk.Label(app, text="Ordem de produção:", bg="lightblue")
    label_op.pack(padx=5, pady=5)

    entry_op = tk.Entry(app, bg="white", bd=2, relief="solid", justify="center")
    entry_op.pack(padx=5, pady=5)

    label_usuario = tk.Label(app, text="Usuário:", bg="lightblue")
    label_usuario.pack(padx=5, pady=5)

    entry_usuario = tk.Entry(app, bg="white", bd=2, relief="solid", justify="center")
    entry_usuario.pack(padx=5, pady=5)

    label_senha = tk.Label(app, text="Senha:",  bg="lightblue")
    label_senha.pack(padx=5, pady=5)

    entry_senha = tk.Entry(app, show="*", bg="white", bd=2, relief="solid", justify="center")
    entry_senha.pack(padx=5, pady=5)

    tk.Button(app, text="Baixar OP's e anexos", command=entrar_lm_e_enviar_para_o_excel, width=20, bg="gold", fg="black").pack(pady=10)

    app.mainloop()

janela = tk.Tk()
janela.title("Projetos Python")
janela.geometry("300x350")
janela.config(bg="lightblue") 

tk.Label(janela, text="Selecione a aplicação desejada", font=("Arial", 14), bg="lightblue").pack(pady=10)
tk.Button(janela, text="Automação de atualização de planilhas", command=atualizar_planilhas_versao_2, width=30, bg="blue", fg="white").pack(pady=5)
tk.Button(janela, text="Automação do sistema nomus", command=atualizar_nomus_versao_1, width=30, bg="blue", fg="white").pack(pady=5)
tk.Button(janela, text="Extração de LM do sistema Nomus", command=extracao_lm_excel, width=30, bg="blue", fg="white").pack(pady=5)
tk.Button(janela, text="Gerar planilhas de ABC", command=abc_geral, width=30, bg="blue", fg="white").pack(pady=5)
tk.Button(janela, text="Gerar planilhas de estoque", command=setores_de_estoque_geral, width=30, bg="blue", fg="white").pack(pady=5)
tk.Button(janela, text="Abrir Calculos", command=calculos_geral, width=30, bg="blue", fg="white").pack(pady=5)
tk.Button(janela, text="Itens importados", command=importados_geral, width=30, bg="blue", fg="white").pack(pady=5)
tk.Button(janela, text="Baixar Ordens e anexos", command=baixar_desenhos, width=30, bg="blue", fg="white").pack(pady=5)

janela.mainloop()
